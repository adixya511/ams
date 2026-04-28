import csv
from datetime import date, datetime
from io import BytesIO, StringIO
# saving the day

from flask import (
    Flask,
    Response,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

app = Flask(__name__)
app.secret_key = "secret123"

# ---------------- DATABASE ----------------
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///attendance.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)


# ---------------- MODELS ----------------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    role = db.Column(db.String(20), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)

    # Student fields
    roll_number = db.Column(db.String(20), nullable=True)
    branch = db.Column(db.String(100), nullable=True)
    semester = db.Column(db.String(20), nullable=True)

    # Teacher fields
    department = db.Column(db.String(100), nullable=True)


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    subject = db.Column(db.String(100), nullable=False)
    branch = db.Column(db.String(100), nullable=True)
    semester = db.Column(db.String(20), nullable=True)

    # added fields required by the templates / routes
    date = db.Column(db.Date, nullable=False, default=date.today)
    status = db.Column(db.String(20), nullable=False, default="absent")

    # optional convenience relationships
    student = db.relationship(
        "User", foreign_keys=[student_id], backref="attendance_as_student"
    )
    teacher = db.relationship(
        "User", foreign_keys=[teacher_id], backref="attendance_as_teacher"
    )


def roll_sort_key(roll_value):
    s = (roll_value or "").strip()
    if s.isdigit():
        return (0, int(s), "")
    return (1, 0, s.lower())


def normalize_subject_name(subject_value):
    s = (subject_value or "").strip()
    if not s:
        return "Unknown"
    if s.lower() == "computer science":
        return "Computer Engineering"
    return s


def subject_match_values(subject_value):
    s = normalize_subject_name(subject_value)
    if s == "Computer Engineering":
        return ["Computer Engineering", "Computer Science"]
    return [s]


TIMETABLE_SUBJECT_ASSIGNMENTS = {
    ("Computer Engineering", "6"): {
        "Talvinder Singh": ["E&SU", "IC"],
        "Avinash Sharma": ["SL", "DCS"],
        "Surbhi Sharma": ["OE-II"],
        "Tamanna": ["OE-III"],
        "Yashwant Singh": ["SCA"],
    }
}

TEACHER_SUBJECT_OPTIONS = [
    "E&SU",
    "Scripting Language",
    "OE 3(multimedia and application)",
    "India constitutuon",
    "OE 2(Data wareahouse and data mining)",
]


def canonicalize_teacher_subject(subject_value):
    subject_raw = (subject_value or "").strip()
    if not subject_raw:
        return ""
    for subject_name in TEACHER_SUBJECT_OPTIONS:
        if subject_raw.lower() == subject_name.lower():
            return subject_name
    return ""


def get_assigned_subjects_for_teacher(teacher, branch=None, semester=None):
    subjects = set()
    if not teacher:
        return []

    # Timetable mapping for class-wise fixed subject assignment.
    if branch and semester:
        class_mapping = TIMETABLE_SUBJECT_ASSIGNMENTS.get(
            ((branch or "").strip(), (semester or "").strip()),
            {},
        )
        for subject_name in class_mapping.get((teacher.name or "").strip(), []):
            subjects.add(normalize_subject_name(subject_name))

    if teacher.department:
        subjects.add(normalize_subject_name(teacher.department))

    query = db.session.query(Attendance.subject).filter(Attendance.teacher_id == teacher.id)
    if branch:
        query = query.filter(Attendance.branch == branch)
    if semester:
        query = query.filter(Attendance.semester == semester)
    for row in query.distinct().all():
        s = normalize_subject_name(row[0])
        if s:
            subjects.add(s)

    return sorted(subjects, key=lambda x: x.lower())


# ---------------- ROOT ----------------
@app.route("/")
def root():
    return redirect(url_for("login"))


# ---------------- LOGIN ----------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        role = request.form.get("role").lower()
        email = request.form.get("email").lower()
        password = request.form.get("password")

        user = User.query.filter(func.lower(User.email) == email).first()

        if (
            not user
            or user.role != role
            or not check_password_hash(user.password, password)
        ):
            flash("Invalid login details!")
            return render_template("login.html")

        session["user_id"] = user.id
        session["role"] = user.role

        if user.role == "admin":
            return redirect(url_for("admin_dashboard"))
        if user.role == "teacher":
            return redirect(url_for("teacher_dashboard"))
        if user.role == "student":
            return redirect(url_for("student_dashboard"))

    return render_template("login.html")


# ---------------- ADMIN ----------------
@app.route("/admin")
def admin_dashboard():
    if session.get("role") != "admin":
        abort(403)

    students = User.query.filter_by(role="student").all()
    teachers = User.query.filter_by(role="teacher").all()
    total_attendance = Attendance.query.count()

    return render_template(
        "admin.html",
        students=students,
        teachers=teachers,
        total_attendance=total_attendance,
    )


@app.route("/admin/students")
def manage_students():
    if session.get("role") != "admin":
        abort(403)

    # Get selected branch and semester from query parameters
    selected_branch = request.args.get("branch", "").strip()
    selected_semester = request.args.get("semester", "").strip()

    # Get all unique branches and semesters from students
    branches = db.session.query(User.branch).filter_by(role="student").distinct().all()
    semesters = (
        db.session.query(User.semester).filter_by(role="student").distinct().all()
    )

    branches = sorted([b[0] for b in branches if b[0]])
    semesters = sorted([s[0] for s in semesters if s[0]])

    # Filter students by branch and/or semester if selected
    query = User.query.filter_by(role="student")
    if selected_branch:
        query = query.filter_by(branch=selected_branch)
    if selected_semester:
        query = query.filter_by(semester=selected_semester)

    students = query.all()

    return render_template(
        "manage_students.html",
        students=students,
        branches=branches,
        semesters=semesters,
        selected_branch=selected_branch,
        selected_semester=selected_semester,
    )


@app.route("/admin/student/add", methods=["GET", "POST"])
def add_student():
    if session.get("role") != "admin":
        abort(403)

    if request.method == "POST":
        u = User(
            name=request.form["name"],
            email=request.form["email"].lower(),
            password=generate_password_hash(request.form["password"]),
            role="student",
            roll_number=request.form["roll_number"],
            branch=request.form.get("branch"),
            semester=request.form.get("semester"),
        )
        db.session.add(u)
        db.session.commit()
        return redirect(url_for("manage_students"))

    # Get unique branches and semesters from existing students
    branches = db.session.query(User.branch).filter_by(role="student").distinct().all()
    semesters = (
        db.session.query(User.semester).filter_by(role="student").distinct().all()
    )

    branches = sorted([b[0] for b in branches if b[0]])
    semesters = sorted([s[0] for s in semesters if s[0]])

    return render_template("add_student.html", branches=branches, semesters=semesters)


@app.route("/admin/student/edit/<int:id>", methods=["GET", "POST"])
def edit_student(id):
    if session.get("role") != "admin":
        abort(403)
    student = User.query.get_or_404(id)
    if request.method == "POST":
        student.name = request.form.get("name", student.name)
        student.email = request.form.get("email", student.email).lower()
        pwd = request.form.get("password", "").strip()
        if pwd:
            student.password = generate_password_hash(pwd)
        student.roll_number = request.form.get("roll_number", student.roll_number)
        student.branch = request.form.get("branch", student.branch)
        student.semester = request.form.get("semester", student.semester)
        db.session.commit()
        flash("Student updated.")
        return redirect(url_for("manage_students"))

    # Get unique branches and semesters from existing students
    branches = db.session.query(User.branch).filter_by(role="student").distinct().all()
    semesters = (
        db.session.query(User.semester).filter_by(role="student").distinct().all()
    )

    branches = sorted([b[0] for b in branches if b[0]])
    semesters = sorted([s[0] for s in semesters if s[0]])

    return render_template(
        "edit_student.html", student=student, branches=branches, semesters=semesters
    )


@app.route("/admin/student/delete/<int:id>")
def delete_student(id):
    if session.get("role") != "admin":
        abort(403)
    db.session.delete(User.query.get_or_404(id))
    db.session.commit()
    flash("Student deleted.")
    return redirect(url_for("manage_students"))


# ---------- TEACHERS ----------
@app.route("/admin/teachers")
def manage_teachers():
    if session.get("role") != "admin":
        abort(403)
    return render_template(
        "manage_teachers.html", teachers=User.query.filter_by(role="teacher").all()
    )


@app.route("/admin/teacher/add", methods=["GET", "POST"])
def add_teacher():
    if session.get("role") != "admin":
        abort(403)

    if request.method == "POST":
        subject = (request.form.get("subject") or "").strip()
        if subject not in TEACHER_SUBJECT_OPTIONS:
            flash("Please select a valid subject for the teacher.")
            semesters = sorted(
                {
                    (row[0] or "").strip()
                    for row in db.session.query(User.semester).filter_by(role="student").distinct().all()
                    if (row[0] or "").strip()
                }
            )
            if not semesters:
                semesters = [str(i) for i in range(1, 9)]
            return render_template(
                "add_teacher.html",
                semesters=semesters,
                subject_options=TEACHER_SUBJECT_OPTIONS,
                selected_subject=subject,
            )
        t = User(
            name=request.form["name"],
            email=request.form["email"].lower(),
            password=generate_password_hash(request.form["password"]),
            role="teacher",
            department=subject,
            semester=(request.form.get("semester") or "").strip() or None,
        )
        db.session.add(t)
        db.session.commit()
        return redirect(url_for("manage_teachers"))

    semesters = sorted(
        {
            (row[0] or "").strip()
            for row in db.session.query(User.semester).filter_by(role="student").distinct().all()
            if (row[0] or "").strip()
        }
    )
    if not semesters:
        semesters = [str(i) for i in range(1, 9)]
    return render_template(
        "add_teacher.html",
        semesters=semesters,
        subject_options=TEACHER_SUBJECT_OPTIONS,
    )


@app.route("/admin/teacher/edit/<int:id>", methods=["GET", "POST"])
def edit_teacher(id):
    if session.get("role") != "admin":
        abort(403)
    teacher = User.query.get_or_404(id)
    if request.method == "POST":
        subject = (request.form.get("subject") or "").strip()
        if subject not in TEACHER_SUBJECT_OPTIONS:
            flash("Please select a valid subject for the teacher.")
            semesters = sorted(
                {
                    (row[0] or "").strip()
                    for row in db.session.query(User.semester).filter_by(role="student").distinct().all()
                    if (row[0] or "").strip()
                }
            )
            if not semesters:
                semesters = [str(i) for i in range(1, 9)]
            return render_template(
                "edit_teacher.html",
                teacher=teacher,
                semesters=semesters,
                subject_options=TEACHER_SUBJECT_OPTIONS,
            )
        teacher.name = request.form.get("name", teacher.name)
        teacher.email = request.form.get("email", teacher.email).lower()
        pwd = request.form.get("password", "").strip()
        if pwd:
            teacher.password = generate_password_hash(pwd)
        teacher.department = subject
        teacher.semester = (request.form.get("semester") or "").strip() or None
        db.session.commit()
        flash("Teacher updated.")
        return redirect(url_for("manage_teachers"))
    semesters = sorted(
        {
            (row[0] or "").strip()
            for row in db.session.query(User.semester).filter_by(role="student").distinct().all()
            if (row[0] or "").strip()
        }
    )
    if not semesters:
        semesters = [str(i) for i in range(1, 9)]
    return render_template(
        "edit_teacher.html",
        teacher=teacher,
        semesters=semesters,
        subject_options=TEACHER_SUBJECT_OPTIONS,
    )


@app.route("/admin/teacher/delete/<int:id>")
def delete_teacher(id):
    if session.get("role") != "admin":
        abort(403)
    db.session.delete(User.query.get_or_404(id))
    db.session.commit()
    flash("Teacher deleted.")
    return redirect(url_for("manage_teachers"))


@app.route("/admin/upload-users-csv", methods=["GET", "POST"])
def upload_users_csv():
    if session.get("role") != "admin":
        abort(403)

    if request.method == "POST":
        f = request.files.get("csv_file")
        if not f or not f.filename:
            flash("Please choose a CSV or Excel file.")
            return redirect(url_for("upload_users_csv"))

        filename = (f.filename or "").strip().lower()
        rows = []
        fieldnames = []

        try:
            file_bytes = f.read()
            if filename.endswith(".csv"):
                content = file_bytes.decode("utf-8-sig", errors="replace")
                reader = csv.DictReader(StringIO(content))
                fieldnames = reader.fieldnames or []
                rows = list(reader)
            elif filename.endswith(".xlsx"):
                if load_workbook is None:
                    flash("Excel support is unavailable. Please install openpyxl or upload CSV.")
                    return redirect(url_for("upload_users_csv"))

                wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
                ws = wb.active
                excel_rows = list(ws.iter_rows(values_only=True))
                if not excel_rows:
                    flash("Uploaded Excel file is empty.")
                    return redirect(url_for("upload_users_csv"))

                fieldnames = [str(h).strip() if h is not None else "" for h in excel_rows[0]]
                rows = []
                for data_row in excel_rows[1:]:
                    row_dict = {}
                    for i, header in enumerate(fieldnames):
                        if not header:
                            continue
                        value = data_row[i] if i < len(data_row) else ""
                        row_dict[header] = "" if value is None else str(value)
                    if any((str(v).strip() for v in row_dict.values())):
                        rows.append(row_dict)
            else:
                flash("Unsupported file type. Please upload .csv or .xlsx file.")
                return redirect(url_for("upload_users_csv"))
        except Exception:
            flash("Unable to read file. Please upload a valid .csv or .xlsx file.")
            return redirect(url_for("upload_users_csv"))

        def _norm_header(h):
            return "".join(ch for ch in (h or "").strip().lower() if ch.isalnum())

        headers = set([h.strip().lower() for h in fieldnames if h])
        norm_headers = set([_norm_header(h) for h in fieldnames if h])
        required_headers = {"name", "password", "email"}
        if not required_headers.issubset(headers) and not required_headers.issubset(norm_headers):
            flash(
                "File must contain headers: name,password,email."
            )
            return redirect(url_for("upload_users_csv"))

        created = 0
        updated = 0
        errors = 0

        for row in rows:
            # Normalize keys so Excel/CSV header case differences don't break parsing.
            normalized_row = {(k or "").strip().lower(): (v or "") for k, v in row.items()}
            normalized_compact_row = {
                _norm_header(k): (v or "") for k, v in row.items() if k
            }

            def get_val(*keys):
                for key in keys:
                    v = normalized_row.get(key)
                    if v not in (None, ""):
                        return str(v).strip()
                for key in keys:
                    v = normalized_compact_row.get(_norm_header(key))
                    if v not in (None, ""):
                        return str(v).strip()
                return ""

            role = get_val("role").lower()
            name = get_val("name")
            email = get_val("email").lower()
            password = get_val("password")
            branch = get_val("branch")
            semester = get_val("semester")
            department = get_val("department")
            roll_number = get_val("roll_number", "roll no", "roll_no", "rollnumber", "roll")

            if not name or not email or not password:
                errors += 1
                continue

            # Auto-detect role when it is not provided.
            if role not in {"student", "teacher"}:
                has_student_fields = bool(branch or semester or roll_number)
                has_teacher_fields = bool(department)
                if has_teacher_fields and not has_student_fields:
                    role = "teacher"
                elif has_student_fields:
                    role = "student"
                else:
                    errors += 1
                    continue

            if role == "student":
                if not branch or not semester:
                    errors += 1
                    continue
            else:
                if not department:
                    errors += 1
                    continue

            existing_user = User.query.filter(func.lower(User.email) == email).first()

            if existing_user:
                existing_user.role = role
                existing_user.name = name
                existing_user.password = generate_password_hash(password)
                existing_user.semester = semester or None

                if role == "student":
                    existing_user.roll_number = (
                        roll_number or None
                    )
                    existing_user.branch = branch or None
                    existing_user.department = None
                else:
                    existing_user.department = department or None
                    existing_user.roll_number = None
                    existing_user.branch = None
                updated += 1
            else:
                if role == "student":
                    u = User(
                        role="student",
                        name=name,
                        email=email,
                        password=generate_password_hash(password),
                        roll_number=roll_number or None,
                        branch=branch or None,
                        semester=semester or None,
                    )
                else:
                    u = User(
                        role="teacher",
                        name=name,
                        email=email,
                        password=generate_password_hash(password),
                        department=department or None,
                        semester=semester or None,
                    )

                db.session.add(u)
                created += 1

        db.session.commit()
        flash(
            f"Upload complete. Created: {created}, Updated(existing): {updated}, Invalid rows: {errors}"
        )
        return redirect(url_for("admin_dashboard"))

    return render_template("upload_users_csv.html")


# ---------------- TEACHER ----------------
@app.route("/teacher")
def teacher_dashboard():
    if session.get("role") != "teacher":
        abort(403)
    teacher = User.query.get(session["user_id"])
    students = User.query.filter_by(role="student").all()
    branches = sorted(
        {
            (row[0] or "").strip()
            for row in db.session.query(User.branch).filter_by(role="student").distinct().all()
            if (row[0] or "").strip()
        }
    )
    semesters = sorted(
        {
            (row[0] or "").strip()
            for row in db.session.query(User.semester).filter_by(role="student").distinct().all()
            if (row[0] or "").strip()
        }
    )
    subjects = list(TEACHER_SUBJECT_OPTIONS)
    return render_template(
        "teacher.html",
        teacher=teacher,
        students=students,
        branches=branches,
        semesters=semesters,
        subjects=subjects,
    )


@app.route("/teacher/quick-action", methods=["POST"])
def teacher_quick_action():
    if session.get("role") != "teacher":
        abort(403)

    branch = (request.form.get("branch") or "").strip()
    semester = (request.form.get("semester") or "").strip()
    subject = canonicalize_teacher_subject(request.form.get("subject"))
    action = (request.form.get("action") or "").strip()

    if not branch or not semester or not subject:
        flash("Please select branch, semester, and subject.")
        return redirect(url_for("teacher_dashboard"))

    if action == "mark":
        session["locked_mark_subject"] = subject
        return redirect(
            url_for("mark_attendance", branch=branch, semester=semester, subject=subject)
        )
    if action == "view":
        return redirect(
            url_for("view_attendance", branch=branch, semester=semester, subject=subject)
        )
    if action == "overall":
        return redirect(
            url_for("overall_attendance", branch=branch, semester=semester, subject=subject)
        )

    flash("Invalid action selected.")
    return redirect(url_for("teacher_dashboard"))


# SELECT BRANCH AND SEMESTER FOR MARKING
@app.route("/teacher/select-class-mark", methods=["GET", "POST"])
def select_class_mark():
    if session.get("role") != "teacher":
        abort(403)

    if request.method == "POST":
        branch = request.form.get("branch")
        semester = request.form.get("semester")
        subject = canonicalize_teacher_subject(request.form.get("subject"))
        if not subject:
            flash("Please select a valid subject.")
            return redirect(url_for("select_class_mark"))
        session["locked_mark_subject"] = subject
        return redirect(
            url_for("mark_attendance", branch=branch, semester=semester, subject=subject)
        )

    # Get unique branches and semesters from students
    branches = db.session.query(User.branch).filter_by(role="student").distinct().all()
    semesters = (
        db.session.query(User.semester).filter_by(role="student").distinct().all()
    )
    subjects = list(TEACHER_SUBJECT_OPTIONS)

    branches = sorted([b[0] for b in branches if b[0]])
    semesters = sorted([s[0] for s in semesters if s[0]])

    return render_template(
        "select_class.html",
        page_title="Select Branch & Semester",
        action="mark_attendance",
        branches=branches,
        semesters=semesters,
        subjects=subjects,
    )


# SELECT BRANCH AND SEMESTER FOR VIEWING
@app.route("/teacher/select-class-view", methods=["GET", "POST"])
def select_class_view():
    if session.get("role") != "teacher":
        abort(403)

    if request.method == "POST":
        branch = request.form.get("branch")
        semester = request.form.get("semester")
        subject = canonicalize_teacher_subject(request.form.get("subject"))
        if not subject:
            flash("Please select a valid subject.")
            return redirect(url_for("select_class_view"))
        return redirect(
            url_for("view_attendance", branch=branch, semester=semester, subject=subject)
        )

    # Get unique branches and semesters from students
    branches = db.session.query(User.branch).filter_by(role="student").distinct().all()
    semesters = (
        db.session.query(User.semester).filter_by(role="student").distinct().all()
    )
    subjects = list(TEACHER_SUBJECT_OPTIONS)

    branches = sorted([b[0] for b in branches if b[0]])
    semesters = sorted([s[0] for s in semesters if s[0]])

    return render_template(
        "select_class.html",
        page_title="Select Branch & Semester",
        action="view_attendance",
        branches=branches,
        semesters=semesters,
        subjects=subjects,
    )


@app.route("/teacher/select-class-overall", methods=["GET", "POST"])
def select_class_overall():
    if session.get("role") != "teacher":
        abort(403)

    if request.method == "POST":
        branch = request.form.get("branch")
        semester = request.form.get("semester")
        subject = canonicalize_teacher_subject(request.form.get("subject"))
        if not subject:
            flash("Please select a valid subject.")
            return redirect(url_for("select_class_overall"))
        return redirect(
            url_for(
                "overall_attendance",
                branch=branch,
                semester=semester,
                subject=subject,
            )
        )

    branches = db.session.query(User.branch).filter_by(role="student").distinct().all()
    semesters = (
        db.session.query(User.semester).filter_by(role="student").distinct().all()
    )
    subjects = list(TEACHER_SUBJECT_OPTIONS)
    branches = sorted([b[0] for b in branches if b[0]])
    semesters = sorted([s[0] for s in semesters if s[0]])

    return render_template(
        "select_class.html",
        page_title="Select Branch & Semester",
        action="overall_attendance",
        branches=branches,
        semesters=semesters,
        subjects=subjects,
    )


@app.route("/teacher/mark", methods=["GET", "POST"])
def mark_attendance():
    if session.get("role") != "teacher":
        abort(403)

    branch = request.args.get("branch")
    semester = request.args.get("semester")
    if not branch or not semester:
        return redirect(url_for("select_class_mark"))

    teacher = User.query.get(session["user_id"])
    requested_subject = canonicalize_teacher_subject(request.args.get("subject"))
    locked_subject = canonicalize_teacher_subject(session.get("locked_mark_subject"))
    if locked_subject and requested_subject and locked_subject != requested_subject:
        flash("Subject changed in URL. Please select subject again.")
        return redirect(url_for("select_class_mark"))
    selected_subject = requested_subject or locked_subject
    if not selected_subject:
        flash("Please select a subject first.")
        return redirect(url_for("select_class_mark"))
    session["locked_mark_subject"] = selected_subject
    students = User.query.filter_by(
        role="student", branch=branch, semester=semester
    ).all()

    selected_date = date.today()
    if request.method == "GET":
        sd = (request.args.get("attendance_date") or "").strip()
        if sd:
            try:
                selected_date = date.fromisoformat(sd)
            except Exception:
                selected_date = date.today()

    existing_status_map = {}
    existing_rows = Attendance.query.filter_by(
        date=selected_date,
        teacher_id=teacher.id,
        subject=selected_subject,
        branch=branch,
        semester=semester,
    ).all()
    for row in existing_rows:
        existing_status_map[row.student_id] = (row.status or "").strip().lower()
    existing_count = len(existing_rows)

    if request.method == "POST":
        sd = (request.form.get("attendance_date") or "").strip()
        if sd:
            try:
                selected_date = date.fromisoformat(sd)
            except Exception:
                selected_date = date.today()

        Attendance.query.filter_by(
            date=selected_date,
            teacher_id=teacher.id,
            subject=selected_subject,
            branch=branch,
            semester=semester,
        ).delete()

        for s in students:
            status = request.form.get(str(s.id))
            if status:
                db.session.add(
                    Attendance(
                        student_id=s.id,
                        teacher_id=teacher.id,
                        subject=selected_subject,
                        branch=branch,
                        semester=semester,
                        date=selected_date,
                        status=status,
                    )
                )

        db.session.commit()
        flash(f"Attendance marked for {selected_date.isoformat()}!")
        return redirect(url_for("teacher_dashboard"))

    return render_template(
        "mark_attendance.html",
        students=students,
        subject=selected_subject,
        teacher=teacher,
        branch=branch,
        semester=semester,
        selected_date=selected_date.isoformat(),
        existing_status_map=existing_status_map,
        existing_count=existing_count,
    )


# ---------------- TEACHER VIEW WITH FILTERS ----------------
@app.route("/teacher/view", methods=["GET", "POST"])
def view_attendance():
    if session.get("role") != "teacher":
        abort(403)

    branch = request.args.get("branch")
    semester = request.args.get("semester")

    if not branch or not semester:
        return redirect(url_for("select_class_view"))

    teacher = User.query.get(session.get("user_id"))

    # teacher's subject is fixed — do not allow choosing other subjects
    assigned_subjects = get_assigned_subjects_for_teacher(teacher, branch, semester)
    selected_subject = (request.args.get("subject") or "").strip() or (
        assigned_subjects[0] if assigned_subjects else None
    )

    # read date range filters
    from_date = None
    to_date = None
    if request.method == "POST":
        fd = (request.form.get("from_date") or "").strip()
        td = (request.form.get("to_date") or "").strip()
        if fd:
            try:
                from_date = date.fromisoformat(fd)
            except Exception:
                from_date = None
        if td:
            try:
                to_date = date.fromisoformat(td)
            except Exception:
                to_date = None
    else:
        fd = request.args.get("from_date", "").strip()
        td = request.args.get("to_date", "").strip()
        if fd:
            try:
                from_date = date.fromisoformat(fd)
            except Exception:
                from_date = None
        if td:
            try:
                to_date = date.fromisoformat(td)
            except Exception:
                to_date = None

    # normalize inverted date range
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date

    # build query — filter by teacher, teacher's subject, branch and semester
    base_query = Attendance.query.filter_by(
        teacher_id=teacher.id, branch=branch, semester=semester
    )
    if selected_subject:
        base_query = base_query.filter_by(subject=selected_subject)

    q = base_query
    if from_date:
        q = q.filter(Attendance.date >= from_date)
    if to_date:
        q = q.filter(Attendance.date <= to_date)

    records = q.order_by(Attendance.date.desc(), Attendance.id.desc()).all()
    for r in records:
        r.student_obj = User.query.get(r.student_id)

    # per-student overall summary for selected range (include all students in class)
    class_students = User.query.filter_by(
        role="student", branch=branch, semester=semester
    ).all()
    counts_by_student = {}
    for r in records:
        c = counts_by_student.setdefault(r.student_id, {"total": 0, "present": 0})
        c["total"] += 1
        if (r.status or "").strip().lower() == "present":
            c["present"] += 1

    student_overall = []
    for s in class_students:
        c = counts_by_student.get(s.id, {"total": 0, "present": 0})
        total = c["total"]
        present = c["present"]
        absent = total - present
        percentage = round((present / total) * 100, 2) if total else 0.0
        student_overall.append(
            {
                "student_id": s.id,
                "name": s.name,
                "roll_number": s.roll_number,
                "email": s.email,
                "present": present,
                "absent": absent,
                "total": total,
                "percentage": percentage,
            }
        )

    student_overall.sort(
        key=lambda x: (roll_sort_key(x["roll_number"]), x["name"].lower())
    )

    # overall summary for selected range
    total_count = q.count()
    present_count = q.filter(func.lower(Attendance.status) == "present").count()
    absent_count = total_count - present_count
    range_percentage = round((present_count / total_count) * 100, 2) if total_count else 0.0
    range_summary = {
        "from_date": from_date.isoformat() if from_date else "",
        "to_date": to_date.isoformat() if to_date else "",
        "total": total_count,
        "present": present_count,
        "absent": absent_count,
        "percentage": range_percentage,
    }

    return render_template(
        "view_attendance.html",
        teacher=teacher,
        records=records,
        from_date=(from_date.isoformat() if from_date else ""),
        to_date=(to_date.isoformat() if to_date else ""),
        range_summary=range_summary,
        student_overall=student_overall,
        selected_subject=selected_subject,
        branch=branch,
        semester=semester,
    )


@app.route("/teacher/overall", methods=["GET", "POST"])
def overall_attendance():
    if session.get("role") != "teacher":
        abort(403)

    branch = request.args.get("branch")
    semester = request.args.get("semester")
    selected_subject = (
        (request.args.get("subject") or request.form.get("subject") or "").strip()
    )
    if not branch or not semester:
        return redirect(url_for("select_class_overall"))

    teacher = User.query.get(session.get("user_id"))

    from_date = None
    to_date = None
    if request.method == "POST":
        fd = (request.form.get("from_date") or "").strip()
        td = (request.form.get("to_date") or "").strip()
    else:
        fd = (request.args.get("from_date") or "").strip()
        td = (request.args.get("to_date") or "").strip()

    if fd:
        try:
            from_date = date.fromisoformat(fd)
        except Exception:
            from_date = None
    if td:
        try:
            to_date = date.fromisoformat(td)
        except Exception:
            to_date = None

    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date

    query = Attendance.query.filter_by(branch=branch, semester=semester)
    if selected_subject:
        query = query.filter(Attendance.subject.in_(subject_match_values(selected_subject)))
    if from_date:
        query = query.filter(Attendance.date >= from_date)
    if to_date:
        query = query.filter(Attendance.date <= to_date)

    records = query.order_by(Attendance.date.desc(), Attendance.id.desc()).all()

    subject_counts = {}
    student_counts = {}
    for r in records:
        subject_key = normalize_subject_name(r.subject)
        subject_stat = subject_counts.setdefault(subject_key, {"total": 0, "present": 0})
        subject_stat["total"] += 1
        if (r.status or "").strip().lower() == "present":
            subject_stat["present"] += 1

        student_stat = student_counts.setdefault(r.student_id, {"total": 0, "present": 0})
        student_stat["total"] += 1
        if (r.status or "").strip().lower() == "present":
            student_stat["present"] += 1

    subject_overall = []
    for subject_name, c in subject_counts.items():
        total = c["total"]
        present = c["present"]
        absent = total - present
        pct = round((present / total) * 100, 2) if total else 0.0
        subject_overall.append(
            {
                "subject": subject_name,
                "present": present,
                "absent": absent,
                "total": total,
                "percentage": pct,
            }
        )
    subject_overall.sort(key=lambda x: x["subject"].lower())

    class_students = User.query.filter_by(
        role="student", branch=branch, semester=semester
    ).all()
    student_overall = []
    for s in class_students:
        c = student_counts.get(s.id, {"total": 0, "present": 0})
        total = c["total"]
        present = c["present"]
        absent = total - present
        pct = round((present / total) * 100, 2) if total else 0.0
        student_overall.append(
            {
                "name": s.name,
                "roll_number": s.roll_number,
                "email": s.email,
                "present": present,
                "absent": absent,
                "total": total,
                "percentage": pct,
            }
        )
    student_overall.sort(
        key=lambda x: (roll_sort_key(x["roll_number"]), x["name"].lower())
    )

    total_count = len(records)
    present_count = sum(
        1 for r in records if (r.status or "").strip().lower() == "present"
    )
    absent_count = total_count - present_count
    overall_percentage = round((present_count / total_count) * 100, 2) if total_count else 0.0

    range_summary = {
        "from_date": from_date.isoformat() if from_date else "",
        "to_date": to_date.isoformat() if to_date else "",
        "total": total_count,
        "present": present_count,
        "absent": absent_count,
        "percentage": overall_percentage,
    }

    return render_template(
        "teacher_overall_attendance.html",
        teacher=teacher,
        branch=branch,
        semester=semester,
        selected_subject=selected_subject,
        from_date=(from_date.isoformat() if from_date else ""),
        to_date=(to_date.isoformat() if to_date else ""),
        range_summary=range_summary,
        subject_overall=subject_overall,
        student_overall=student_overall,
    )


@app.route("/teacher/download-csv")
def download_attendance_csv():
    if session.get("role") != "teacher":
        abort(403)

    teacher = User.query.get(session.get("user_id"))

    branch = request.args.get("branch")
    semester = request.args.get("semester")
    assigned_subjects = get_assigned_subjects_for_teacher(teacher, branch, semester)
    selected_subject = (request.args.get("subject") or "").strip() or (
        assigned_subjects[0] if assigned_subjects else None
    )
    from_date_filter = request.args.get("from_date")
    to_date_filter = request.args.get("to_date")

    if not branch or not semester:
        abort(400)

    query = Attendance.query.filter_by(
        teacher_id=teacher.id,
        branch=branch,
        semester=semester,
        subject=selected_subject,
    )

    if from_date_filter:
        try:
            query = query.filter(Attendance.date >= date.fromisoformat(from_date_filter))
        except:
            pass
    if to_date_filter:
        try:
            query = query.filter(Attendance.date <= date.fromisoformat(to_date_filter))
        except:
            pass

    records = query.order_by(Attendance.date.desc()).all()

    # per-student overall summary for selected range (include all students in class)
    class_students = User.query.filter_by(
        role="student", branch=branch, semester=semester
    ).all()
    counts_by_student = {}
    for r in records:
        c = counts_by_student.setdefault(r.student_id, {"total": 0, "present": 0})
        c["total"] += 1
        if (r.status or "").strip().lower() == "present":
            c["present"] += 1

    si = StringIO()
    cw = csv.writer(si)

    cw.writerow(["Per Student Overall Attendance"])
    cw.writerow(["Student Name", "Roll Number", "Present", "Absent", "Total", "Overall %"])
    for s in sorted(class_students, key=lambda x: (roll_sort_key(x.roll_number), x.name.lower())):
        c = counts_by_student.get(s.id, {"total": 0, "present": 0})
        total = c["total"]
        present = c["present"]
        absent = total - present
        pct = round((present / total) * 100, 2) if total else 0.0
        cw.writerow([s.name, s.roll_number, present, absent, total, f"{pct}%"])

    output = si.getvalue()
    si.close()

    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=attendance.csv"},
    )


@app.route("/teacher/overall/download-csv")
def download_overall_attendance_csv():
    if session.get("role") != "teacher":
        abort(403)

    branch = request.args.get("branch")
    semester = request.args.get("semester")
    selected_subject = (request.args.get("subject") or "").strip()
    from_date_filter = (request.args.get("from_date") or "").strip()
    to_date_filter = (request.args.get("to_date") or "").strip()

    if not branch or not semester:
        abort(400)

    from_date = None
    to_date = None
    if from_date_filter:
        try:
            from_date = date.fromisoformat(from_date_filter)
        except Exception:
            from_date = None
    if to_date_filter:
        try:
            to_date = date.fromisoformat(to_date_filter)
        except Exception:
            to_date = None
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date

    query = Attendance.query.filter_by(branch=branch, semester=semester)
    if selected_subject:
        query = query.filter(Attendance.subject.in_(subject_match_values(selected_subject)))
    if from_date:
        query = query.filter(Attendance.date >= from_date)
    if to_date:
        query = query.filter(Attendance.date <= to_date)

    records = query.all()

    student_counts = {}
    for r in records:
        sc = student_counts.setdefault(r.student_id, {"total": 0, "present": 0})
        sc["total"] += 1
        if (r.status or "").strip().lower() == "present":
            sc["present"] += 1

    class_students = User.query.filter_by(
        role="student", branch=branch, semester=semester
    ).all()

    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(["Student-wise Overall Attendance (All Subjects)"])
    cw.writerow(["Student", "Roll Number", "Email", "Present", "Absent", "Total", "Overall %"])

    for s in sorted(class_students, key=lambda x: (roll_sort_key(x.roll_number), x.name.lower())):
        c = student_counts.get(s.id, {"total": 0, "present": 0})
        total = c["total"]
        present = c["present"]
        absent = total - present
        pct = round((present / total) * 100, 2) if total else 0.0
        cw.writerow([s.name, s.roll_number or "", s.email, present, absent, total, f"{pct}%"])

    output = si.getvalue()
    si.close()

    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=student_overall_attendance.csv"},
    )


# ---------------- STUDENT ----------------
@app.route("/student", methods=["GET", "POST"])
def student_dashboard():
    if session.get("role") != "student":
        abort(403)

    student_id = session["user_id"]
    student = User.query.get(student_id)

    teachers = User.query.filter_by(role="teacher").all()
    selected_branch = (request.form.get("branch") or "").strip()
    selected_date = request.form.get("selected_date")

    query = Attendance.query.filter_by(student_id=student_id)

    if selected_branch:
        query = query.filter(Attendance.branch == selected_branch)

    if selected_date:
        selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        query = query.filter_by(date=selected_date)

    records = query.all()
    for r in records:
        # Keep DB-backed field immutable in request scope to avoid autoflush updates.
        r.display_subject = normalize_subject_name(r.subject)

    # Filtered stats
    total_days = len(records)
    present_days = sum(1 for r in records if r.status.lower() == "present")
    absent_days = total_days - present_days
    percentage = round((present_days / total_days) * 100, 2) if total_days > 0 else 0

    # ------------------------------------------
    # 1. OVERALL ATTENDANCE (All subjects, all dates)
    # ------------------------------------------
    all_records = Attendance.query.filter_by(student_id=student_id).all()
    all_total = len(all_records)
    all_present = sum(1 for r in all_records if r.status.lower() == "present")
    overall_percentage = (
        round((all_present / all_total) * 100, 2) if all_total > 0 else 0
    )

    # ------------------------------------------
    # 2. MONTHLY ATTENDANCE
    # ------------------------------------------
    today = date.today()
    first_day = today.replace(day=1)

    monthly_records = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.date >= first_day,
        Attendance.date <= today,
    ).all()

    monthly_total = len(monthly_records)
    monthly_present = sum(1 for r in monthly_records if r.status.lower() == "present")
    monthly_percentage = (
        round((monthly_present / monthly_total) * 100, 2) if monthly_total > 0 else 0
    )

    today_record = Attendance.query.filter_by(
        student_id=student_id, date=date.today()
    ).first()
    branches = sorted(
        {
            (r[0] or "").strip()
            for r in db.session.query(Attendance.branch)
            .filter(Attendance.student_id == student_id)
            .distinct()
            .all()
            if (r[0] or "").strip()
        },
        key=lambda x: x.lower(),
    )

    return render_template(
        "student.html",
        student=student,
        teachers=teachers,
        selected_branch=selected_branch,
        selected_date=selected_date,
        records=records,
        total_days=total_days,
        present_days=present_days,
        absent_days=absent_days,
        percentage=percentage,
        today_record=today_record,
        branches=branches,
        overall_percentage=overall_percentage,
        monthly_percentage=monthly_percentage,
    )


@app.route("/student/change-password", methods=["GET", "POST"])
def student_change_password():
    if session.get("role") != "student":
        abort(403)

    student = User.query.get(session.get("user_id"))
    if not student:
        abort(404)

    if request.method == "POST":
        current_password = (request.form.get("current_password") or "").strip()
        new_password = (request.form.get("new_password") or "").strip()
        confirm_password = (request.form.get("confirm_password") or "").strip()

        if not current_password or not new_password or not confirm_password:
            flash("Please fill all password fields.")
            return render_template("student_change_password.html", student=student)

        if not check_password_hash(student.password, current_password):
            flash("Current password is incorrect.")
            return render_template("student_change_password.html", student=student)

        if len(new_password) < 6:
            flash("New password must be at least 6 characters.")
            return render_template("student_change_password.html", student=student)

        if new_password != confirm_password:
            flash("New password and confirm password do not match.")
            return render_template("student_change_password.html", student=student)

        if check_password_hash(student.password, new_password):
            flash("New password cannot be the same as current password.")
            return render_template("student_change_password.html", student=student)

        student.password = generate_password_hash(new_password)
        db.session.commit()
        flash("Password changed successfully.")
        return redirect(url_for("student_change_password"))

    return render_template("student_change_password.html", student=student)


@app.route("/student/subjects")
def student_subjects():
    if session.get("role") != "student":
        abort(403)

    student = User.query.get(session.get("user_id"))
    if not student:
        abort(404)

    # --- branches where this student has any record ---
    records_for_student = Attendance.query.filter_by(student_id=student.id).all()
    branches = sorted(
        {(r.branch or "").strip() for r in records_for_student if (r.branch or "").strip()},
        key=lambda x: x.lower(),
    )

    fallback = False

    # fallback: if student has no records, show all branches in system
    if not branches:
        rows_all = db.session.query(Attendance.branch).distinct().all()
        branches = sorted(
            {(r[0] or "").strip() for r in rows_all if (r[0] or "").strip()},
            key=lambda x: x.lower(),
        )
        fallback = True

    subjects_info = []

    for branch_name in branches:
        # total distinct dates (classes)
        total_sessions = (
            db.session.query(func.count(func.distinct(Attendance.date)))
            .filter(Attendance.branch == branch_name)
            .scalar()
            or 0
        )

        present_count = (
            db.session.query(Attendance)
            .filter(
                Attendance.branch == branch_name,
                Attendance.student_id == student.id,
                func.lower(Attendance.status) == "present",
            )
            .count()
        )

        # last class date
        last_rec = (
            db.session.query(Attendance.date)
            .filter(Attendance.branch == branch_name)
            .order_by(Attendance.date.desc())
            .first()
        )
        last_date = last_rec[0].isoformat() if last_rec else None

        # teacher names for this branch
        teacher_rows = (
            db.session.query(User.name)
            .join(Attendance, Attendance.teacher_id == User.id)
            .filter(Attendance.branch == branch_name)
            .distinct()
            .all()
        )
        teacher_names = ", ".join([t[0] for t in teacher_rows]) if teacher_rows else ""

        # calculate percentage
        percent = (
            round((present_count / total_sessions) * 100, 1)
            if total_sessions > 0
            else 0
        )

        # check if student has any attendance in this branch
        enrolled = (
            Attendance.query.filter(
                Attendance.branch == branch_name,
                Attendance.student_id == student.id,
            ).count()
            > 0
        )

        subjects_info.append(
            {
                "branch": branch_name,
                "total_sessions": total_sessions,
                "present_count": present_count,
                "present_percent": percent,
                "last_date": last_date,
                "teachers": teacher_names,
                "enrolled": enrolled,
            }
        )

    return render_template(
        "student_subjects.html",
        student=student,
        subjects_info=subjects_info,
        fallback=fallback,
    )


@app.route("/student/subject/<path:subject>/details")
def student_subject_details(subject):
    if session.get("role") != "student":
        abort(403)
    student_id = session.get("user_id")

    normalized_subject = normalize_subject_name(subject)
    match_values = subject_match_values(normalized_subject)

    # total distinct session dates for this subject (class-wide)
    total_sessions = (
        db.session.query(func.count(func.distinct(Attendance.date)))
        .filter(Attendance.subject.in_(match_values))
        .scalar()
        or 0
    )

    # count present for this student
    present_count = Attendance.query.filter(
        Attendance.subject.in_(match_values),
        Attendance.student_id == student_id,
        func.lower(Attendance.status) == "present",
    ).count()

    # per-session history for this student (date + status)
    hist_rows = (
        db.session.query(Attendance.date, Attendance.status)
        .filter(
            Attendance.subject.in_(match_values),
            Attendance.student_id == student_id,
        )
        .order_by(Attendance.date)
        .all()
    )
    history = [{"date": r[0].isoformat(), "status": r[1]} for r in hist_rows]

    return jsonify(
        {
            "subject": normalized_subject,
            "total_sessions": total_sessions,
            "present_count": present_count,
            "present_percent": (
                round((present_count / total_sessions) * 100, 1)
                if total_sessions > 0
                else 0.0
            ),
            "history": history,
        }
    )


@app.route("/student/classes")
def student_classes():
    if session.get("role") != "student":
        abort(403)
    student = User.query.get(session.get("user_id"))
    # build per-branch attendance stats for this student
    records_for_student = Attendance.query.filter_by(student_id=student.id).all()
    branches = sorted(
        {(r.branch or "").strip() for r in records_for_student if (r.branch or "").strip()},
        key=lambda x: x.lower(),
    )
    classes_stats = []
    for branch_name in branches:
        total = Attendance.query.filter(
            Attendance.student_id == student.id,
            Attendance.branch == branch_name,
        ).count()
        present = Attendance.query.filter(
            Attendance.student_id == student.id,
            Attendance.branch == branch_name,
            func.lower(Attendance.status) == "present",
        ).count()
        pct = round((present / total) * 100, 2) if total > 0 else 0
        classes_stats.append(
            {"branch": branch_name, "total": total, "present": present, "percentage": pct}
        )
    return render_template(
        "student_classes.html", student=student, classes_stats=classes_stats
    )


@app.route("/student/attendance")
def student_attendance():
    if session.get("role") != "student":
        abort(403)
    student = User.query.get(session.get("user_id"))
    records = (
        Attendance.query.filter_by(student_id=student.id)
        .order_by(Attendance.date.desc())
        .all()
    )
    for r in records:
        r.teacher_obj = User.query.get(r.teacher_id)
        r.display_subject = normalize_subject_name(r.subject)
    return render_template("student_attendance.html", student=student, records=records)


# ---------------- REPORTS ----------------
@app.route("/reports")
def reports():
    if session.get("role") != "admin":
        abort(403)

    selected_branch = (request.args.get("branch") or "").strip()
    selected_semester = (request.args.get("semester") or "").strip()

    # options for filter dropdowns
    branches = {
        (row[0] or "").strip()
        for row in db.session.query(User.branch).filter_by(role="student").distinct().all()
        if row[0]
    }
    branches.update(
        {
            (row[0] or "").strip()
            for row in db.session.query(Attendance.branch).distinct().all()
            if row[0]
        }
    )
    semesters = {
        (row[0] or "").strip()
        for row in db.session.query(User.semester).filter_by(role="student").distinct().all()
        if row[0]
    }
    semesters.update(
        {
            (row[0] or "").strip()
            for row in db.session.query(Attendance.semester).distinct().all()
            if row[0]
        }
    )
    branch_options = sorted(branches)
    semester_options = sorted(semesters)

    # filtered attendance base query
    attendance_query = Attendance.query
    if selected_branch:
        attendance_query = attendance_query.filter(Attendance.branch == selected_branch)
    if selected_semester:
        attendance_query = attendance_query.filter(Attendance.semester == selected_semester)

    records_all = attendance_query.all()
    total_attendance = len(records_all)

    # filtered students count
    students_query = User.query.filter_by(role="student")
    if selected_branch:
        students_query = students_query.filter(User.branch == selected_branch)
    if selected_semester:
        students_query = students_query.filter(User.semester == selected_semester)
    students = students_query.all()
    total_students = len(students)

    # filtered teachers count (teachers with attendance in current filter)
    teacher_ids = sorted({r.teacher_id for r in records_all})
    total_teachers = len(teacher_ids)

    # Attendance by subject
    grouped_subjects = {}
    for r in records_all:
        key = normalize_subject_name(r.subject)
        c = grouped_subjects.setdefault(key, {"total": 0, "present": 0})
        c["total"] += 1
        if (r.status or "").strip().lower() == "present":
            c["present"] += 1

    attendance_by_subject = []
    for subj in sorted(grouped_subjects.keys(), key=lambda x: x.lower()):
        total = grouped_subjects[subj]["total"]
        present = grouped_subjects[subj]["present"]
        pct = round((present / total) * 100, 2) if total > 0 else 0
        attendance_by_subject.append(
            {"subject": subj, "total": total, "present": present, "percentage": pct}
        )

    # Attendance by teacher
    teachers = (
        User.query.filter(User.role == "teacher", User.id.in_(teacher_ids)).all()
        if teacher_ids
        else []
    )
    attendance_by_teacher = []
    for t in teachers:
        total = sum(1 for r in records_all if r.teacher_id == t.id)
        present = sum(
            1
            for r in records_all
            if r.teacher_id == t.id and (r.status or "").strip().lower() == "present"
        )
        pct = round((present / total) * 100, 2) if total > 0 else 0
        attendance_by_teacher.append(
            {"teacher": t.name, "total": total, "present": present, "percentage": pct}
        )
    attendance_by_teacher.sort(key=lambda x: x["teacher"].lower())

    # Students with lowest attendance (overall) - top 10
    student_stats = []
    for s in students:
        total = sum(1 for r in records_all if r.student_id == s.id)
        present = sum(
            1
            for r in records_all
            if r.student_id == s.id and (r.status or "").strip().lower() == "present"
        )
        pct = round((present / total) * 100, 2) if total > 0 else None
        student_stats.append(
            {
                "id": s.id,
                "name": s.name,
                "email": s.email,
                "total": total,
                "present": present,
                "percentage": pct,
            }
        )
    low_attendance = sorted(
        [x for x in student_stats if x["percentage"] is not None],
        key=lambda r: r["percentage"],
    )[:10]

    # Recent attendance entries
    recent_records = (
        attendance_query.order_by(Attendance.date.desc(), Attendance.id.desc())
        .limit(12)
        .all()
    )
    # attach student & teacher names for display
    for r in recent_records:
        r.student_obj = User.query.get(r.student_id)
        r.teacher_obj = User.query.get(r.teacher_id)
        r.display_subject = normalize_subject_name(r.subject)

    return render_template(
        "reports.html",
        total_students=total_students,
        total_teachers=total_teachers,
        total_attendance=total_attendance,
        attendance_by_subject=attendance_by_subject,
        attendance_by_teacher=attendance_by_teacher,
        low_attendance=low_attendance,
        recent_records=recent_records,
        branch_options=branch_options,
        semester_options=semester_options,
        selected_branch=selected_branch,
        selected_semester=selected_semester,
    )


# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------- TEACHER CLASSES ----------------
@app.route("/teacher/classes")
def teacher_classes():
    if session.get("role") != "teacher":
        abort(403)
    teacher = User.query.get(session.get("user_id"))

    # build unique class list (teacher.department + subjects from attendance)
    classes = []
    if teacher and teacher.department:
        classes.append(normalize_subject_name(teacher.department))
    if teacher:
        subjects = [
            row[0]
            for row in db.session.query(Attendance.subject)
            .filter_by(teacher_id=teacher.id)
            .distinct()
            .all()
            or []
        ]
        for s in subjects:
            normalized = normalize_subject_name(s)
            if normalized and normalized not in classes:
                classes.append(normalized)

    # build richer info for each class (sessions count, distinct students, last session date)
    classes_info = []
    for s in classes:
        match_values = subject_match_values(s)
        total_sessions = Attendance.query.filter_by(
            teacher_id=teacher.id
        ).filter(Attendance.subject.in_(match_values)
        ).count()
        last_rec = (
            Attendance.query.filter_by(teacher_id=teacher.id)
            .filter(Attendance.subject.in_(match_values))
            .order_by(Attendance.date.desc())
            .first()
        )
        last_date = last_rec.date.isoformat() if last_rec else None
        students_count = (
            db.session.query(Attendance.student_id)
            .filter_by(teacher_id=teacher.id)
            .filter(Attendance.subject.in_(match_values))
            .distinct()
            .count()
        )
        classes_info.append(
            {
                "subject": s,
                "total_sessions": total_sessions,
                "last_date": last_date,
                "students_count": students_count,
            }
        )

    return render_template(
        "teacher_classes.html", teacher=teacher, classes_info=classes_info
    )


# ---------------- UTILITY FUNCTIONS ----------------
def get_all_subjects():
    subs = set()
    # from attendance table
    rows = db.session.query(Attendance.subject).distinct().all()
    for r in rows:
        s = normalize_subject_name(r[0])
        if s:
            subs.add(s)
    return sorted(subs)


# ---------------- MAIN ----------------
if __name__ == "__main__":
    with app.app_context():
        db.create_all()

        # ONLY CREATE DEFAULT USERS IF DB IS EMPTY
        if User.query.count() == 0:
            admin = User(
                name="Admin",
                email="admin@gmail.com",
                password=generate_password_hash("admin123"),
                role="admin",
            )
            default_teacher = User(
                name="Mr. Sharma",
                email="teacher@gmail.com",
                password=generate_password_hash("teacher123"),
                role="teacher",
                department="CS",
            )

            # Additional teacher/staff teaching accounts
            # password kept common for first login; they can change later from profile/admin edit
            teachers_data = [
                (
                    "Naresh Kumar",
                    "nareshkumarspch@gmail.com",
                    "CS",
                    "Computer Engineering",
                ),
                (
                    "Rajesh Sharma",
                    "rsharma72@gmail.com",
                    "Electrical",
                    "Electrical Engineering",
                ),
                (
                    "Hari Singh Thakur",
                    "hsthakur09@gmail.com",
                    "ECE",
                    "Electronics Engineering",
                ),
                ("Kamlesh Chand", "kchand78@gmail.com", "AS&H", "Applied Science"),
                (
                    "Pawan Chandel",
                    "pawanchandel13@gmail.com",
                    "Instrumentation",
                    "Instrumentation Engineering",
                ),
                ("Talvinder Singh", "talvinder.m@gmail.com", "CS", "Computer Engineering"),
                ("Jagdeep Singh", "jagdeep9906@gmail.com", "CS", "Computer Engineering"),
                (
                    "Sudhir Dhiman",
                    "sudhir.dhiman85@gmail.com",
                    "CS",
                    "Computer Engineering",
                ),
                ("Onkar Singh", "onkar.singh26970@gmail.com", "CS", "Computer Engineering"),
                ("Saroop Chand", "saroop2388@gmail.com", "CS", "Computer Engineering"),
                (
                    "Sachin Sehota",
                    "sachin.sehota@gmail.com",
                    "ECE",
                    "Electronics Engineering",
                ),
                ("Rohit Kumar", "rohit06d@gmail.com", "ECE", "Electronics Engineering"),
                (
                    "Nishant Kaushal",
                    "nishant.1987@gmail.com",
                    "ECE",
                    "Electronics Engineering",
                ),
                (
                    "Raman Kumar",
                    "ramankumar89@gmail.com",
                    "ECE",
                    "Electronics Engineering",
                ),
                (
                    "Rajeev Kumar",
                    "rajeev.kumar357@gmail.com",
                    "CE",
                    "Computer Engineering",
                ),
                ("Avinash Sharma", "avinash.acet@yahoo.com", "CE", "Computer Engineering"),
                (
                    "Surbhi Sharma",
                    "surbhisharma.jmj@gmail.com",
                    "CE",
                    "Computer Engineering",
                ),
                ("Tamanna", "er.tamanna14@gmail.com", "CE", "Computer Engineering"),
                ("Yashwant Singh", "yashwant.singh.ce@gmail.com", "CE", "Computer Engineering"),
                ("Ashima Sharma", "ashima143188@gmail.com", "CE", "Computer Engineering"),
                (
                    "Vijay Kumar Sharma",
                    "vijay2122@gmail.com",
                    "CE",
                    "Computer Engineering",
                ),
                (
                    "Vikal Sharma",
                    "vikal.in.sharma@gmail.com",
                    "IE",
                    "Instrumentation Engineering",
                ),
                (
                    "Karan Singh Thakur",
                    "karansingh.thakur89@gmail.com",
                    "IE",
                    "Instrumentation Engineering",
                ),
                ("Varun", "varunkrr72@gmail.com", "IE", "Instrumentation Engineering"),
                (
                    "Ritika Sharma",
                    "ritika30purohit30@gmail.com",
                    "IE",
                    "Instrumentation Engineering",
                ),
                (
                    "Munish Kumar",
                    "munishsharma198@gmail.com",
                    "IE",
                    "Instrumentation Engineering",
                ),
                (
                    "Santosh Kumar",
                    "kumarsantosh9606@gmail.com",
                    "ME",
                    "Mechanical Engineering",
                ),
                (
                    "Mohan Lal",
                    "mohanpathania6747@gmail.com",
                    "ME",
                    "Mechanical Engineering",
                ),
                (
                    "Subash Chand",
                    "swastiksneha@gmail.com",
                    "ME",
                    "Mechanical Engineering",
                ),
                (
                    "Satbir Singh",
                    "satbirsuru@gmail.com",
                    "ME",
                    "Mechanical Engineering",
                ),
                (
                    "Sameer Sharma",
                    "sameer30132@gmail.com",
                    "ME",
                    "Mechanical Engineering",
                ),
                ("Ina Gupta", "guptaina24@gmail.com", "EE", "Electrical Engineering"),
                ("Iela Bharti", "iela.b89@gmail.com", "EE", "Electrical Engineering"),
                ("Pritam Chand", "pritam777018@gmail.com", "Physics", "Physics"),
                ("Richa Sharma", "richathakur121@gmail.com", "English", "English"),
                ("Vijay Thakur", "thakurviju454@gmail.com", "Physics", "Physics"),
                ("Anil Kumar", "prashanil77@gmail.com", "Chemistry", "Chemistry"),
            ]

            # Create students data
            students_data = [
                # Computer Engineering - 2nd Semester (47 students)
                ("Aarkit Sharma", "aarkit@gmail.com", "1", "Computer Engineering", "2"),
                ("Abansik", "abansik@gmail.com", "2", "Computer Engineering", "2"),
                ("Abhay Dogra", "abhay.dogra@gmail.com", "3", "Computer Engineering", "2"),
                ("Abhay Verma", "abhay.verma@gmail.com", "4", "Computer Engineering", "2"),
                ("Aditya", "aditya@gmail.com", "5", "Computer Engineering", "2"),
                ("Akshit Dev", "akshit@gmail.com", "6", "Computer Engineering", "2"),
                ("Anmol Saklani", "anmol@gmail.com", "7", "Computer Engineering", "2"),
                ("Anshil", "anshil@gmail.com", "8", "Computer Engineering", "2"),
                ("Anshuman", "anshuman@gmail.com", "9", "Computer Engineering", "2"),
                ("Arun Kumar", "arun.kumar@gmail.com", "10", "Computer Engineering", "2"),
                ("Arush Patiyal", "arush@gmail.com", "11", "Computer Engineering", "2"),
                ("Aryan", "aryan@gmail.com", "12", "Computer Engineering", "2"),
                ("Ayush", "ayush@gmail.com", "13", "Computer Engineering", "2"),
                ("Divyansh", "divyansh@gmail.com", "14", "Computer Engineering", "2"),
                ("Hemant Koundal", "hemant@gmail.com", "15", "Computer Engineering", "2"),
                ("Harshit", "harshit@gmail.com", "16", "Computer Engineering", "2"),
                ("Ishan Sagar", "ishan@gmail.com", "17", "Computer Engineering", "2"),
                ("Isharit Dogra", "isharit@gmail.com", "18", "Computer Engineering", "2"),
                ("Ishita Sugha", "ishita@gmail.com", "19", "Computer Engineering", "2"),
                ("Kajal", "kajal@gmail.com", "20", "Computer Engineering", "2"),
                ("Krish Bharti", "krish@gmail.com", "21", "Computer Engineering", "2"),
                ("Manan Sharma", "manan@gmail.com", "22", "Computer Engineering", "2"),
                ("Manisha Dhiman", "manisha@gmail.com", "23", "Computer Engineering", "2"),
                ("Mayank Baggae", "mayank@gmail.com", "24", "Computer Engineering", "2"),
                ("Muskan", "muskan@gmail.com", "25", "Computer Engineering", "2"),
                ("Nitsh Kumar", "nitsh@gmail.com", "26", "Computer Engineering", "2"),
                ("Paras", "paras@gmail.com", "27", "Computer Engineering", "2"),
                ("Radhika", "radhika@gmail.com", "28", "Computer Engineering", "2"),
                ("Raj Sharma", "raj@gmail.com", "29", "Computer Engineering", "2"),
                (
                    "Ridhima Choudhary",
                    "ridhima@gmail.com",
                    "30",
                    "Computer Engineering",
                    "2",
                ),
                (
                    "Rudrasish Singh Rana",
                    "rudrasish@gmail.com",
                    "31",
                    "Computer Engineering",
                    "2",
                ),
                ("Sahil Guleria", "sahil@gmail.com", "32", "Computer Engineering", "2"),
                ("Sanjeev Kumar", "sanjeev@gmail.com", "33", "Computer Engineering", "2"),
                ("Sanvi Mahajan", "sanvi@gmail.com", "34", "Computer Engineering", "2"),
                ("Sheetal", "sheetal@gmail.com", "35", "Computer Engineering", "2"),
                ("Shivam", "shivam@gmail.com", "36", "Computer Engineering", "2"),
                (
                    "Shivam Sharma",
                    "shivam.sharma@gmail.com",
                    "37",
                    "Computer Engineering",
                    "2",
                ),
                (
                    "Shivam Thakur",
                    "shivam.thakur@gmail.com",
                    "38",
                    "Computer Engineering",
                    "2",
                ),
                ("Shivend Gautam", "shivend@gmail.com", "39", "Computer Engineering", "2"),
                ("Simran Choudhary", "simran@gmail.com", "40", "Computer Engineering", "2"),
                ("Sohani Dogra", "sohani@gmail.com", "41", "Computer Engineering", "2"),
                (
                    "Surya Kumar Koundal",
                    "surya@gmail.com",
                    "42",
                    "Computer Engineering",
                    "2",
                ),
                ("Swastik Naryal", "swastik@gmail.com", "43", "Computer Engineering", "2"),
                ("Tamanna Sharma", "tamanna@gmail.com", "44", "Computer Engineering", "2"),
                (
                    "Utkarsh Choudhary",
                    "utkarsh@gmail.com",
                    "45",
                    "Computer Engineering",
                    "2",
                ),
                ("Yashita Mehra", "yashita@gmail.com", "46", "Computer Engineering", "2"),
                ("Yogesh Raj", "yogesh@gmail.com", "47", "Computer Engineering", "2"),
                # Computer Engineering - 4th Semester (52 students)
                ("Aarshit Rana", "aarshit@gmail.com", "1", "Computer Engineering", "4"),
                ("Aavya Parmar", "aavya@gmail.com", "2", "Computer Engineering", "4"),
                ("Aayush Parmar", "aayush@gmail.com", "3", "Computer Engineering", "4"),
                ("Abhishek Chalib", "abhishek@gmail.com", "4", "Computer Engineering", "4"),
                (
                    "Aditya Sharma",
                    "aditya.sharma@gmail.com",
                    "5",
                    "Computer Engineering",
                    "4",
                ),
                ("Akash", "akash@gmail.com", "6", "Computer Engineering", "4"),
                ("Akhil Sharma", "akhil@gmail.com", "7", "Computer Engineering", "4"),
                ("Amartya", "amartya@gmail.com", "8", "Computer Engineering", "4"),
                ("Ankush", "ankush@gmail.com", "9", "Computer Engineering", "4"),
                (
                    "Antriksh Choudhary",
                    "antriksh@gmail.com",
                    "10",
                    "Computer Engineering",
                    "4",
                ),
                ("Anush Kumar", "anush@gmail.com", "11", "Computer Engineering", "4"),
                ("Anushaka", "anushaka@gmail.com", "12", "Computer Engineering", "4"),
                ("Anushka Bhatta", "anushka@gmail.com", "13", "Computer Engineering", "4"),
                ("Apoorva Dogra", "apoorva@gmail.com", "14", "Computer Engineering", "4"),
                ("Arfan", "arfan@gmail.com", "15", "Computer Engineering", "4"),
                (
                    "Ayush Thakur",
                    "ayush.thakur@gmail.com",
                    "16",
                    "Computer Engineering",
                    "4",
                ),
                (
                    "Bhuvaneshwar",
                    "bhuvaneshwar@gmail.com",
                    "17",
                    "Computer Engineering",
                    "4",
                ),
                (
                    "Chaitanya Sharma",
                    "chaitanya@gmail.com",
                    "18",
                    "Computer Engineering",
                    "4",
                ),
                ("Hardev Singh", "hardev@gmail.com", "19", "Computer Engineering", "4"),
                ("Harish Mondga", "harish@gmail.com", "20", "Computer Engineering", "4"),
                ("Janvi", "janvi@gmail.com", "21", "Computer Engineering", "4"),
                ("Karun Shiva", "karun@gmail.com", "22", "Computer Engineering", "4"),
                ("Kavya", "kavya@gmail.com", "23", "Computer Engineering", "4"),
                ("Kriti Rana", "kriti@gmail.com", "24", "Computer Engineering", "4"),
                ("Lakshay Sharma", "lakshay@gmail.com", "25", "Computer Engineering", "4"),
                ("Manish", "manish@gmail.com", "26", "Computer Engineering", "4"),
                (
                    "Maninder Uldeen",
                    "maninder@gmail.com",
                    "27",
                    "Computer Engineering",
                    "4",
                ),
                ("Manvi Rana", "manvi@gmail.com", "28", "Computer Engineering", "4"),
                ("Mohit Bharti", "mohit@gmail.com", "29", "Computer Engineering", "4"),
                ("Mridul", "mridul@gmail.com", "30", "Computer Engineering", "4"),
                ("Muskan", "muskan.cs4@gmail.com", "31", "Computer Engineering", "4"),
                ("Nikhil Choudhary", "nikhil@gmail.com", "32", "Computer Engineering", "4"),
                ("Nishant Kumar", "nishant@gmail.com", "33", "Computer Engineering", "4"),
                ("Palak Bhangale", "palak@gmail.com", "34", "Computer Engineering", "4"),
                ("Ridhim Kumar", "ridhim@gmail.com", "35", "Computer Engineering", "4"),
                ("Rishabh Kaundal", "rishabh@gmail.com", "36", "Computer Engineering", "4"),
                (
                    "Shrishti Sharma",
                    "shrishti@gmail.com",
                    "37",
                    "Computer Engineering",
                    "4",
                ),
                ("Sneha", "sneha@gmail.com", "38", "Computer Engineering", "4"),
                ("Suhaan Mehra", "suhaan@gmail.com", "39", "Computer Engineering", "4"),
                ("Tanvi Sharma", "tanvi@gmail.com", "40", "Computer Engineering", "4"),
                ("Uday Singh", "uday@gmail.com", "41", "Computer Engineering", "4"),
                ("Vipul", "vipul@gmail.com", "42", "Computer Engineering", "4"),
                ("Anshika", "anshika@gmail.com", "43", "Computer Engineering", "4"),
                ("Jeeya Patiyal", "jeeya@gmail.com", "44", "Computer Engineering", "4"),
                ("Kamlesh Kumari", "kamlesh@gmail.com", "45", "Computer Engineering", "4"),
                (
                    "Manish Kumar",
                    "manish.kumar@gmail.com",
                    "46",
                    "Computer Engineering",
                    "4",
                ),
                ("Diksha", "diksha@gmail.com", "47", "Computer Engineering", "4"),
                ("Harshit Rai", "harshit.rai@gmail.com", "48", "Computer Engineering", "4"),
                ("Ashish", "ashish@gmail.com", "49", "Computer Engineering", "4"),
                ("Nalin Koundal", "nalin@gmail.com", "50", "Computer Engineering", "4"),
                ("Ankita Kumari", "ankita@gmail.com", "51", "Computer Engineering", "4"),
                ("Dhruv", "dhruv@gmail.com", "52", "Computer Engineering", "4"),
                # Computer Engineering - 6th Semester (50 students)
                (
                    "Aarush Koundal",
                    "aarush.koundal@gmail.com",
                    "1",
                    "Computer Engineering",
                    "6",
                ),
                ("Aditya", "aditya.6@gmail.com", "2", "Computer Engineering", "6"),
                ("Aditya", "aditya.6b@gmail.com", "3", "Computer Engineering", "6"),
                (
                    "Aditya Thakur",
                    "aditya.thakur@gmail.com",
                    "4",
                    "Computer Engineering",
                    "6",
                ),
                (
                    "Aditya Katoch",
                    "aditya.katoch@gmail.com",
                    "5",
                    "Computer Engineering",
                    "6",
                ),
                (
                    "Aditya Sharma",
                    "aditya.sharma.6@gmail.com",
                    "6",
                    "Computer Engineering",
                    "6",
                ),
                ("Akasshi Mehra", "akasshi@gmail.com", "7", "Computer Engineering", "6"),
                ("Akshara Thakur", "akshara@gmail.com", "8", "Computer Engineering", "6"),
                ("Arkshit Sharma", "arkshit@gmail.com", "9", "Computer Engineering", "6"),
                ("Anita", "anita@gmail.com", "10", "Computer Engineering", "6"),
                ("Areen", "areen@gmail.com", "11", "Computer Engineering", "6"),
                (
                    "Aryan Dhiman",
                    "aryan.dhiman@gmail.com",
                    "12",
                    "Computer Engineering",
                    "6",
                ),
                (
                    "Aryan Jamwal",
                    "aryan.jamwal@gmail.com",
                    "13",
                    "Computer Engineering",
                    "6",
                ),
                ("Ayush", "ayush.6@gmail.com", "14", "Computer Engineering", "6"),
                ("Banshul Kumar", "banshul@gmail.com", "15", "Computer Engineering", "6"),
                (
                    "Diksha Chauhan",
                    "diksha.chauhan@gmail.com",
                    "16",
                    "Computer Engineering",
                    "6",
                ),
                ("Divyanshi", "divyanshi@gmail.com", "17", "Computer Engineering", "6"),
                ("Harsh", "harsh@gmail.com", "18", "Computer Engineering", "6"),
                ("Harshi", "harshi@gmail.com", "19", "Computer Engineering", "6"),
                (
                    "Harshit Kapoor",
                    "harshit.kapoor@gmail.com",
                    "20",
                    "Computer Engineering",
                    "6",
                ),
                ("Isha Kumari", "isha.kumari@gmail.com", "21", "Computer Engineering", "6"),
                ("Ishaan Kumar", "ishaan@gmail.com", "22", "Computer Engineering", "6"),
                (
                    "Muskan Choudhary",
                    "muskan.choudhary@gmail.com",
                    "23",
                    "Computer Engineering",
                    "6",
                ),
                ("Piyush", "piyush@gmail.com", "24", "Computer Engineering", "6"),
                ("Priya", "priya@gmail.com", "25", "Computer Engineering", "6"),
                (
                    "Pushkar Pathania",
                    "pushkar@gmail.com",
                    "26",
                    "Computer Engineering",
                    "6",
                ),
                ("Rasik Jarwal", "rasik@gmail.com", "27", "Computer Engineering", "6"),
                ("Riya Dhiman", "riya.dhiman@gmail.com", "28", "Computer Engineering", "6"),
                ("Sarshi Koundal", "sarshi@gmail.com", "29", "Computer Engineering", "6"),
                ("Saniya", "saniya@gmail.com", "30", "Computer Engineering", "6"),
                ("Sejal", "sejal@gmail.com", "31", "Computer Engineering", "6"),
                ("Shagun", "shagun@gmail.com", "32", "Computer Engineering", "6"),
                ("Shiven Sharma", "shiven@gmail.com", "33", "Computer Engineering", "6"),
                ("Shreya", "shreya@gmail.com", "34", "Computer Engineering", "6"),
                (
                    "Simran Bharti",
                    "simran.bharti@gmail.com",
                    "35",
                    "Computer Engineering",
                    "6",
                ),
                ("Sneha", "sneha.6@gmail.com", "36", "Computer Engineering", "6"),
                ("Tanisha", "tanisha@gmail.com", "37", "Computer Engineering", "6"),
                (
                    "Vanshika Naryal",
                    "vanshika@gmail.com",
                    "38",
                    "Computer Engineering",
                    "6",
                ),
                (
                    "Vishakha Koundal",
                    "vishakha@gmail.com",
                    "39",
                    "Computer Engineering",
                    "6",
                ),
                (
                    "Ankit Koundal",
                    "ankit.koundal@gmail.com",
                    "40",
                    "Computer Engineering",
                    "6",
                ),
                (
                    "Ankit Gulerja",
                    "ankit.gulerja@gmail.com",
                    "41",
                    "Computer Engineering",
                    "6",
                ),
                ("Akshay Kumar", "akshay@gmail.com", "42", "Computer Engineering", "6"),
                ("Abhihay Thakur", "abhihay@gmail.com", "43", "Computer Engineering", "6"),
                (
                    "Parshant Sharma",
                    "parshant@gmail.com",
                    "44",
                    "Computer Engineering",
                    "6",
                ),
                ("Uday Thakur", "uday.thakur@gmail.com", "45", "Computer Engineering", "6"),
                ("Anshika", "anshika.6@gmail.com", "46", "Computer Engineering", "6"),
                ("Vansh Rana", "vansh@gmail.com", "47", "Computer Engineering", "6"),
                (
                    "Disha Choudhary",
                    "disha.choudhary@gmail.com",
                    "48",
                    "Computer Engineering",
                    "6",
                ),
                ("Mannat Walia", "mannat@gmail.com", "49", "Computer Engineering", "6"),
                (
                    "Krish Singh Athwal",
                    "krish.singh@gmail.com",
                    "50",
                    "Computer Engineering",
                    "6",
                ),
                # Electrical Engineering - 2nd Semester (45 students)
                (
                    "Abhishek Mittal",
                    "abhishek.mittal@gmail.com",
                    "1",
                    "Electrical",
                    "2",
                ),
                ("Aditya Nanda", "aditya.nanda@gmail.com", "2", "Electrical", "2"),
                ("Ankit Karnoot", "ankit.karnoot@gmail.com", "3", "Electrical", "2"),
                ("Ankit", "ankit@gmail.com", "4", "Electrical", "2"),
                ("Avinash Devi", "avinash.devi@gmail.com", "5", "Electrical", "2"),
                ("Adyoti Choudhary", "adyoti@gmail.com", "6", "Electrical", "2"),
                (
                    "Amit Kumar Maharaj Singh",
                    "amit.kumar@gmail.com",
                    "7",
                    "Electrical",
                    "2",
                ),
                ("Amata Harjps", "amata@gmail.com", "8", "Electrical", "2"),
                ("Amu Sharma", "amu@gmail.com", "9", "Electrical", "2"),
                ("Anurag Mahtra", "anurag@gmail.com", "10", "Electrical", "2"),
                ("Bhuvan Dev", "bhuvan@gmail.com", "11", "Electrical", "2"),
                ("Dighansh", "dighansh@gmail.com", "12", "Electrical", "2"),
                ("Harshul Thakpur", "harshul@gmail.com", "13", "Electrical", "2"),
                ("Harshir Choudhary", "harshir@gmail.com", "14", "Electrical", "2"),
                ("Ishant", "ishant@gmail.com", "15", "Electrical", "2"),
                ("Jaswant Kaur", "jaswant@gmail.com", "16", "Electrical", "2"),
                ("Kartik", "kartik@gmail.com", "17", "Electrical", "2"),
                (
                    "Kartik Choudhary",
                    "kartik.choudhary@gmail.com",
                    "18",
                    "Electrical",
                    "2",
                ),
                ("Kartik Kumar", "kartik.kumar@gmail.com", "19", "Electrical", "2"),
                ("Karvey", "karvey@gmail.com", "20", "Electrical", "2"),
                ("Kishor Sharma", "kishor@gmail.com", "21", "Electrical", "2"),
                ("Kiriti Kumar", "kiriti@gmail.com", "22", "Electrical", "2"),
                ("Nitesh Harpal", "nitesh@gmail.com", "23", "Electrical", "2"),
                ("Pallab Sharma", "pallab@gmail.com", "24", "Electrical", "2"),
                ("Rebindai", "rebindai@gmail.com", "25", "Electrical", "2"),
                ("Resnav Kumar", "resnav@gmail.com", "26", "Electrical", "2"),
                (
                    "Rishi Choudhary",
                    "rishi.choudhary@gmail.com",
                    "27",
                    "Electrical",
                    "2",
                ),
                ("Rishi", "rishi@gmail.com", "28", "Electrical", "2"),
                (
                    "Rohit Choudhary",
                    "rohit.choudhary@gmail.com",
                    "29",
                    "Electrical",
                    "2",
                ),
                ("Rohit Ola Rana", "rohit.rana@gmail.com", "30", "Electrical", "2"),
                ("Sahid Kumar", "sahid@gmail.com", "31", "Electrical", "2"),
                ("Sameet Sharma", "sameet@gmail.com", "32", "Electrical", "2"),
                ("Satvirram Daitnam", "satvirram@gmail.com", "33", "Electrical", "2"),
                ("Satyam Singh", "satyam@gmail.com", "34", "Electrical", "2"),
                ("Shardham Jintyal", "shardham@gmail.com", "35", "Electrical", "2"),
                ("Shridham", "shridham@gmail.com", "36", "Electrical", "2"),
                (
                    "Simran Choudhary",
                    "simran.choudhary@gmail.com",
                    "37",
                    "Electrical",
                    "2",
                ),
                ("Shalun Sharma", "shalun@gmail.com", "38", "Electrical", "2"),
                ("Sumit Kunbal", "sumit@gmail.com", "39", "Electrical", "2"),
                ("Taksh Batshai", "taksh@gmail.com", "40", "Electrical", "2"),
                ("Tanish Koushal", "tanish@gmail.com", "41", "Electrical", "2"),
                ("Vikram Chantyal", "vikram@gmail.com", "42", "Electrical", "2"),
                ("Abhinav", "abhinav@gmail.com", "43", "Electrical", "2"),
                ("Akarsht Choudhary", "akarsht@gmail.com", "44", "Electrical", "2"),
                ("Ileshan Kalihal", "ileshan@gmail.com", "45", "Electrical", "2"),
                # Electrical - 4th Semester (49 students)
                ("Abhishek Koli", "abhishek.koli@gmail.com", "1", "Electrical", "4"),
                ("Kanish Sharma", "kanish@gmail.com", "2", "Electrical", "4"),
                ("Abishek Jintyal", "abishek@gmail.com", "3", "Electrical", "4"),
                ("Aditya", "aditya.e4@gmail.com", "4", "Electrical", "4"),
                ("Aditya Katoch", "aditya.katoch.e@gmail.com", "5", "Electrical", "4"),
                ("Aditya Dhiman", "aditya.dhiman@gmail.com", "6", "Electrical", "4"),
                ("Akhil", "akhil.e@gmail.com", "7", "Electrical", "4"),
                ("Akhil Kumar", "akhil.kumar.e@gmail.com", "8", "Electrical", "4"),
                ("Akshata", "akshata@gmail.com", "9", "Electrical", "4"),
                ("Akshata Devi", "akshata.devi@gmail.com", "10", "Electrical", "4"),
                ("Anmol", "anmol.e@gmail.com", "11", "Electrical", "4"),
                ("Anuruddha", "anuruddha@gmail.com", "12", "Electrical", "4"),
                ("Apurva", "apurva@gmail.com", "13", "Electrical", "4"),
                ("Arjun", "arjun@gmail.com", "14", "Electrical", "4"),
                (
                    "Arjun Choudhary",
                    "arjun.choudhary@gmail.com",
                    "15",
                    "Electrical",
                    "4",
                ),
                ("Aryush Kumar", "aryush@gmail.com", "16", "Electrical", "4"),
                ("Ashfaq", "ashfaq@gmail.com", "17", "Electrical", "4"),
                ("Ashok", "ashok@gmail.com", "18", "Electrical", "4"),
                ("Ashok Dogra", "ashok.dogra@gmail.com", "19", "Electrical", "4"),
                ("Aswat", "aswat@gmail.com", "20", "Electrical", "4"),
                ("Aswini", "aswini@gmail.com", "21", "Electrical", "4"),
                (
                    "Aswini Choudhary",
                    "aswini.choudhary@gmail.com",
                    "22",
                    "Electrical",
                    "4",
                ),
                ("Atri Din", "atri@gmail.com", "23", "Electrical", "4"),
                ("Avnish", "avnish@gmail.com", "24", "Electrical", "4"),
                ("Avnish Kumar", "avnish.kumar@gmail.com", "25", "Electrical", "4"),
                ("Baljit", "baljit@gmail.com", "26", "Electrical", "4"),
                ("Baljit Nath", "baljit.nath@gmail.com", "27", "Electrical", "4"),
                ("Basant", "basant@gmail.com", "28", "Electrical", "4"),
                ("Bhagwan Prasad", "bhagwan@gmail.com", "29", "Electrical", "4"),
                ("Bhavesh Sangha", "bhavesh@gmail.com", "30", "Electrical", "4"),
                ("Bhola", "bhola@gmail.com", "31", "Electrical", "4"),
                ("Bhuma", "bhuma@gmail.com", "32", "Electrical", "4"),
                ("Bhupendra", "bhupendra@gmail.com", "33", "Electrical", "4"),
                (
                    "Bikram Choudhary",
                    "bikram.choudhary@gmail.com",
                    "34",
                    "Electrical",
                    "4",
                ),
                ("Bikram Sharma", "bikram.sharma@gmail.com", "35", "Electrical", "4"),
                ("Brajesh Singh Guleria", "brajesh@gmail.com", "36", "Electrical", "4"),
                ("Britannia Choudhary", "britannia@gmail.com", "37", "Electrical", "4"),
                ("Chirash Mahajan", "chirash@gmail.com", "38", "Electrical", "4"),
                ("Chaman", "chaman@gmail.com", "39", "Electrical", "4"),
                ("Chamkur Singh", "chamkur@gmail.com", "40", "Electrical", "4"),
                ("Chet Ram", "chet@gmail.com", "41", "Electrical", "4"),
                ("Chirag", "chirag@gmail.com", "42", "Electrical", "4"),
                ("Chiranji", "chiranji@gmail.com", "43", "Electrical", "4"),
                ("Devinder Singh", "devinder@gmail.com", "44", "Electrical", "4"),
                ("Deepak", "deepak@gmail.com", "45", "Electrical", "4"),
                ("Dewan", "dewan@gmail.com", "46", "Electrical", "4"),
                ("Dharamvir Singh", "dharamvir@gmail.com", "47", "Electrical", "4"),
                ("Dhiraj", "dhiraj@gmail.com", "48", "Electrical", "4"),
                ("Yogesh", "yogesh.e@gmail.com", "49", "Electrical", "4"),
                # Electrical - 6th Semester (52 students)
                ("Aadi Chauhan", "aadi@gmail.com", "1", "Electrical", "6"),
                ("Abhishek", "abhishek.e6@gmail.com", "2", "Electrical", "6"),
                ("Abhishek", "abhishek.e6b@gmail.com", "3", "Electrical", "6"),
                ("Akhil Kumar", "akhil.e6@gmail.com", "4", "Electrical", "6"),
                ("Akshay Sharma", "akshay.sharma@gmail.com", "5", "Electrical", "6"),
                ("Amit Kumar", "amit.e6@gmail.com", "6", "Electrical", "6"),
                ("Anchal Kumar", "anchal@gmail.com", "7", "Electrical", "6"),
                ("Aniket", "aniket@gmail.com", "8", "Electrical", "6"),
                ("Ankur Verma", "ankur@gmail.com", "9", "Electrical", "6"),
                ("Anshul", "anshul.e6@gmail.com", "10", "Electrical", "6"),
                ("Anshul", "anshul.e6b@gmail.com", "11", "Electrical", "6"),
                ("Anju Kumar", "anju.e6@gmail.com", "12", "Electrical", "6"),
                ("Aryan Sharma", "aryan.sharma@gmail.com", "13", "Electrical", "6"),
                ("Aastha Thakur", "aastha@gmail.com", "14", "Electrical", "6"),
                ("Avinit", "avinit@gmail.com", "15", "Electrical", "6"),
                ("Ayush Koundal", "ayush.koundal@gmail.com", "16", "Electrical", "6"),
                ("Bhuprendra Singh", "bhuprendra@gmail.com", "17", "Electrical", "6"),
                ("Disha Choudhary", "disha@gmail.com", "18", "Electrical", "6"),
                ("Gourav Koundal", "gourav@gmail.com", "19", "Electrical", "6"),
                ("Harsh Sharma", "harsh.sharma@gmail.com", "20", "Electrical", "6"),
                (
                    "Harshit Choudhary",
                    "harshit.choudhary@gmail.com",
                    "21",
                    "Electrical",
                    "6",
                ),
                ("Himanshu Choudhary", "himanshu@gmail.com", "22", "Electrical", "6"),
                ("Ishant", "ishant.e6@gmail.com", "23", "Electrical", "6"),
                ("Karishthik", "karishthik@gmail.com", "24", "Electrical", "6"),
                ("Kartik Choudhary", "kartik.e6@gmail.com", "25", "Electrical", "6"),
                ("Kartik Thakur", "kartik.thakur@gmail.com", "26", "Electrical", "6"),
                ("Kshitij Goswan", "kshitij@gmail.com", "27", "Electrical", "6"),
                ("Kushav Rana", "kushav@gmail.com", "28", "Electrical", "6"),
                ("Mansi", "mansi@gmail.com", "29", "Electrical", "6"),
                ("Hussain Choudhary", "hussain@gmail.com", "30", "Electrical", "6"),
                ("Pratham Chandel", "pratham@gmail.com", "31", "Electrical", "6"),
                ("Praveen", "praveen.e6@gmail.com", "32", "Electrical", "6"),
                (
                    "Sahil Choudhary",
                    "sahil.choudhary@gmail.com",
                    "33",
                    "Electrical",
                    "6",
                ),
                ("Sahil Kumar", "sahil.kumar@gmail.com", "34", "Electrical", "6"),
                ("Sandeep Kumar", "sandeep.e6@gmail.com", "35", "Electrical", "6"),
                ("Sanjeev Kumar", "sanjeev.e6@gmail.com", "36", "Electrical", "6"),
                ("Shunny", "shunny@gmail.com", "37", "Electrical", "6"),
                ("Dikshant Rana", "dikshant@gmail.com", "38", "Electrical", "6"),
                ("Swasthik", "swasthik.e6@gmail.com", "39", "Electrical", "6"),
                ("Varun Kumar", "varun.e6@gmail.com", "40", "Electrical", "6"),
                (
                    "Abhishek Kumar",
                    "abhishek.kumar.e@gmail.com",
                    "41",
                    "Electrical",
                    "6",
                ),
                ("Kritika", "kritika@gmail.com", "42", "Electrical", "6"),
                ("Palak", "palak.e6@gmail.com", "43", "Electrical", "6"),
                ("Sejal", "sejal.e6@gmail.com", "44", "Electrical", "6"),
                ("Sujal Choudhary", "sujal@gmail.com", "45", "Electrical", "6"),
                ("Akshay", "akshay.e6@gmail.com", "46", "Electrical", "6"),
                ("Akshit Kumar", "akshit.e6@gmail.com", "47", "Electrical", "6"),
                ("Rahul Dhiman", "rahul.dhiman@gmail.com", "48", "Electrical", "6"),
                ("Banshul Ratra", "banshul.e6@gmail.com", "49", "Electrical", "6"),
                ("Aashit Rana", "aashit@gmail.com", "50", "Electrical", "6"),
                ("Priyanshu", "priyanshu@gmail.com", "51", "Electrical", "6"),
                ("Gautam Choudhary", "gautam@gmail.com", "52", "Electrical", "6"),
                # Electronics - 4th Semester (41 students)
                ("Abhav Choudhary", "abhav@gmail.com", "1", "Electronics", "4"),
                ("Aditya Verma", "aditya.verma.ec@gmail.com", "2", "Electronics", "4"),
                ("Aditya Verma", "aditya.verma.ec2@gmail.com", "3", "Electronics", "4"),
                ("Akhil Rana", "akhil.rana.ec@gmail.com", "4", "Electronics", "4"),
                (
                    "Akshay Choudhary",
                    "akshay.choudhary.ec@gmail.com",
                    "5",
                    "Electronics",
                    "4",
                ),
                ("Akshita", "akshita.ec@gmail.com", "6", "Electronics", "4"),
                ("Akshita Kumari", "akshita.kumari@gmail.com", "7", "Electronics", "4"),
                ("Aletia", "aletia@gmail.com", "8", "Electronics", "4"),
                ("Amar", "amar.ec@gmail.com", "9", "Electronics", "4"),
                ("Anish Kumar", "anish.kumar.ec@gmail.com", "10", "Electronics", "4"),
                ("Anikit", "anikit@gmail.com", "11", "Electronics", "4"),
                ("Amanpreet", "amanpreet@gmail.com", "12", "Electronics", "4"),
                (
                    "Amar Singh Chauhan",
                    "amar.singh@gmail.com",
                    "13",
                    "Electronics",
                    "4",
                ),
                ("Ashish", "ashish.ec@gmail.com", "14", "Electronics", "4"),
                ("Ankush", "ankush.ec@gmail.com", "15", "Electronics", "4"),
                ("Anshul", "anshul.ec@gmail.com", "16", "Electronics", "4"),
                ("Koushal Kumar", "koushal@gmail.com", "17", "Electronics", "4"),
                ("Kuman Verma", "kuman@gmail.com", "18", "Electronics", "4"),
                (
                    "Nikhil Koundal",
                    "nikhil.koundal.ec@gmail.com",
                    "19",
                    "Electronics",
                    "4",
                ),
                ("Rituesh Bharti", "rituesh@gmail.com", "20", "Electronics", "4"),
                ("Rishab", "rishab@gmail.com", "21", "Electronics", "4"),
                ("Priyanshu", "priyanshu.ec@gmail.com", "22", "Electronics", "4"),
                ("Pulkit Desai", "pulkit@gmail.com", "23", "Electronics", "4"),
                ("Barshil Saklani", "barshil@gmail.com", "24", "Electronics", "4"),
                ("Riya Nanda", "riya.nanda@gmail.com", "25", "Electronics", "4"),
                ("Sanjeep", "sanjeep@gmail.com", "26", "Electronics", "4"),
                ("Saicsar Koundal", "saicsar@gmail.com", "27", "Electronics", "4"),
                ("Shabaev Sharma", "shabaev@gmail.com", "28", "Electronics", "4"),
                ("Shubham", "shubham.ec@gmail.com", "29", "Electronics", "4"),
                ("Sunan Choudhary", "sunan@gmail.com", "30", "Electronics", "4"),
                ("Tanishk Guptai", "tanishk@gmail.com", "31", "Electronics", "4"),
                ("Taraishi Sharma", "taraishi@gmail.com", "32", "Electronics", "4"),
                ("Tashvir Singh", "tashvir@gmail.com", "33", "Electronics", "4"),
                ("Vishal", "vishal.ec@gmail.com", "34", "Electronics", "4"),
                ("Vishnesh Sharma", "vishnesh@gmail.com", "35", "Electronics", "4"),
                ("Anshul", "anshul.ec2@gmail.com", "36", "Electronics", "4"),
                ("Anuj", "anuj.ec@gmail.com", "37", "Electronics", "4"),
                ("Arjun Sharma", "arjun.sharma.ec@gmail.com", "38", "Electronics", "4"),
                ("Krish", "krish.ec@gmail.com", "39", "Electronics", "4"),
                ("Kuk Kum", "kuk@gmail.com", "40", "Electronics", "4"),
                ("Sakshi", "sakshi@gmail.com", "41", "Electronics", "4"),
                ("Shivani", "shivani@gmail.com", "42", "Electronics", "4"),
                ("Aryan", "aryan.ec@gmail.com", "43", "Electronics", "4"),
                ("Sahil", "sahil.ec@gmail.com", "44", "Electronics", "4"),
                # Electronics - 6th Semester (38 students)
                ("Abhishek", "abhishek.ec6@gmail.com", "1", "Electronics", "6"),
                ("Abidheye Choudhary", "abidheye@gmail.com", "2", "Electronics", "6"),
                ("Akshet Dhiman", "akshet@gmail.com", "3", "Electronics", "6"),
                ("Ankit Dawahal", "ankit.dawahal@gmail.com", "4", "Electronics", "6"),
                ("Areshita Thakur", "areshita@gmail.com", "5", "Electronics", "6"),
                ("Ashvi", "ashvi@gmail.com", "6", "Electronics", "6"),
                ("Bhuwan Singh", "bhuwan.singh@gmail.com", "7", "Electronics", "6"),
                (
                    "Gaurav Choudhary",
                    "gaurav.choudhary@gmail.com",
                    "8",
                    "Electronics",
                    "6",
                ),
                (
                    "Himanshu Choudhary",
                    "himanshu.choudhary@gmail.com",
                    "9",
                    "Electronics",
                    "6",
                ),
                ("Luv Kumar", "luv@gmail.com", "10", "Electronics", "6"),
                (
                    "Mahak Choudhary",
                    "mahak.choudhary@gmail.com",
                    "11",
                    "Electronics",
                    "6",
                ),
                ("Mehak Dogra", "mehak@gmail.com", "12", "Electronics", "6"),
                ("Nidhi", "nidhi@gmail.com", "13", "Electronics", "6"),
                (
                    "Nitka Choudhary",
                    "nitka.choudhary@gmail.com",
                    "14",
                    "Electronics",
                    "6",
                ),
                ("Nitka Kumari", "nitka.kumari@gmail.com", "15", "Electronics", "6"),
                ("Nitin", "nitin@gmail.com", "16", "Electronics", "6"),
                ("Parihit Sharma", "parihit@gmail.com", "17", "Electronics", "6"),
                ("Raghita Thakur", "raghita@gmail.com", "18", "Electronics", "6"),
                ("Ritika", "ritika@gmail.com", "19", "Electronics", "6"),
                ("Rustam", "rustam@gmail.com", "20", "Electronics", "6"),
                ("Sakshi Devi", "sakshi.devi@gmail.com", "21", "Electronics", "6"),
                ("Sameck Sharma", "sameck@gmail.com", "22", "Electronics", "6"),
                ("Baluni", "baluni@gmail.com", "23", "Electronics", "6"),
                ("Shweta Kumari", "shweta@gmail.com", "24", "Electronics", "6"),
                ("Suneha", "suneha@gmail.com", "25", "Electronics", "6"),
                ("Tanisha", "tanisha.ec6@gmail.com", "26", "Electronics", "6"),
                ("Tanishka", "tanishka@gmail.com", "27", "Electronics", "6"),
                ("Varun Guleria", "varun.guleria@gmail.com", "28", "Electronics", "6"),
                (
                    "Ayush Choudhary",
                    "ayush.choudhary.ec6@gmail.com",
                    "29",
                    "Electronics",
                    "6",
                ),
                ("Baneshita", "baneshita@gmail.com", "30", "Electronics", "6"),
                ("Kartik Bhatt", "kartik.bhatt@gmail.com", "31", "Electronics", "6"),
                ("Kunti Sharma", "kunti@gmail.com", "32", "Electronics", "6"),
                ("Laxman Thakur", "laxman@gmail.com", "33", "Electronics", "6"),
                ("Riya Kumari", "riya.kumari.ec@gmail.com", "34", "Electronics", "6"),
                ("Ritik Sharma", "ritik.sharma@gmail.com", "35", "Electronics", "6"),
                (
                    "Aditya Sharma",
                    "aditya.sharma.ec6@gmail.com",
                    "36",
                    "Electronics",
                    "6",
                ),
                (
                    "Govind Singh",
                    "govind.singh.ec6@gmail.com",
                    "37",
                    "Electronics",
                    "6",
                ),
                ("Dikshant Rana", "dikshant.rana@gmail.com", "38", "Electronics", "6"),
                # Electronics - 2nd Semester (50 students)
                (
                    "Abhishek Mittal",
                    "abhishek.mittal.ec2@gmail.com",
                    "1",
                    "Electronics",
                    "2",
                ),
                (
                    "Aditya Verma",
                    "aditya.verma.ec2a@gmail.com",
                    "2",
                    "Electronics",
                    "2",
                ),
                (
                    "Aditya Verma",
                    "aditya.verma.ec2b@gmail.com",
                    "3",
                    "Electronics",
                    "2",
                ),
                (
                    "Aditya Thakur",
                    "aditya.thakur.ec@gmail.com",
                    "4",
                    "Electronics",
                    "2",
                ),
                (
                    "Aditya Katoch",
                    "aditya.katoch.ec@gmail.com",
                    "5",
                    "Electronics",
                    "2",
                ),
                (
                    "Aditya Sharma",
                    "aditya.sharma.ec2@gmail.com",
                    "6",
                    "Electronics",
                    "2",
                ),
                ("Akasshi Mehra", "akasshi.mehra@gmail.com", "7", "Electronics", "2"),
                ("Akshara Thakur", "akshara.thakur@gmail.com", "8", "Electronics", "2"),
                ("Arkshit Sharma", "arkshit.sharma@gmail.com", "9", "Electronics", "2"),
                ("Anita", "anita.ec@gmail.com", "10", "Electronics", "2"),
                ("Areen", "areen.ec@gmail.com", "11", "Electronics", "2"),
                ("Aryan Dhiman", "aryan.dhiman.ec@gmail.com", "12", "Electronics", "2"),
                ("Aryan Jamwal", "aryan.jamwal.ec@gmail.com", "13", "Electronics", "2"),
                ("Ayush", "ayush.ec@gmail.com", "14", "Electronics", "2"),
                (
                    "Banshul Kumar",
                    "banshul.kumar.ec@gmail.com",
                    "15",
                    "Electronics",
                    "2",
                ),
                (
                    "Diksha Chauhan",
                    "diksha.chauhan.ec@gmail.com",
                    "16",
                    "Electronics",
                    "2",
                ),
                ("Divyanshi", "divyanshi.ec@gmail.com", "17", "Electronics", "2"),
                ("Harsh", "harsh.ec@gmail.com", "18", "Electronics", "2"),
                ("Harshi", "harshi.ec@gmail.com", "19", "Electronics", "2"),
                (
                    "Harshit Kapoor",
                    "harshit.kapoor.ec@gmail.com",
                    "20",
                    "Electronics",
                    "2",
                ),
                ("Isha Kumari", "isha.kumari.ec@gmail.com", "21", "Electronics", "2"),
                ("Ishaan Kumar", "ishaan.kumar.ec@gmail.com", "22", "Electronics", "2"),
                (
                    "Muskan Choudhary",
                    "muskan.choudhary.ec@gmail.com",
                    "23",
                    "Electronics",
                    "2",
                ),
                ("Piyush", "piyush.ec@gmail.com", "24", "Electronics", "2"),
                ("Priya", "priya.ec@gmail.com", "25", "Electronics", "2"),
                (
                    "Pushkar Pathania",
                    "pushkar.pathania@gmail.com",
                    "26",
                    "Electronics",
                    "2",
                ),
                ("Rasik Jarwal", "rasik.jarwal@gmail.com", "27", "Electronics", "2"),
                ("Riya Dhiman", "riya.dhiman.ec@gmail.com", "28", "Electronics", "2"),
                (
                    "Sarshi Koundal",
                    "sarshi.koundal@gmail.com",
                    "29",
                    "Electronics",
                    "2",
                ),
                ("Saniya", "saniya.ec@gmail.com", "30", "Electronics", "2"),
                ("Sejal", "sejal.ec@gmail.com", "31", "Electronics", "2"),
                ("Shagun", "shagun.ec@gmail.com", "32", "Electronics", "2"),
                (
                    "Shiven Sharma",
                    "shiven.sharma.ec@gmail.com",
                    "33",
                    "Electronics",
                    "2",
                ),
                ("Shreya", "shreya.ec@gmail.com", "34", "Electronics", "2"),
                (
                    "Simran Bharti",
                    "simran.bharti.ec@gmail.com",
                    "35",
                    "Electronics",
                    "2",
                ),
                ("Sneha", "sneha.ec@gmail.com", "36", "Electronics", "2"),
                ("Tanisha", "tanisha.ec2@gmail.com", "37", "Electronics", "2"),
                (
                    "Vanshika Naryal",
                    "vanshika.naryal@gmail.com",
                    "38",
                    "Electronics",
                    "2",
                ),
                (
                    "Vishakha Koundal",
                    "vishakha.koundal.ec@gmail.com",
                    "39",
                    "Electronics",
                    "2",
                ),
                (
                    "Ankit Koundal",
                    "ankit.koundal.ec@gmail.com",
                    "40",
                    "Electronics",
                    "2",
                ),
                (
                    "Ankit Gulerja",
                    "ankit.gulerja.ec@gmail.com",
                    "41",
                    "Electronics",
                    "2",
                ),
                ("Akshay Kumar", "akshay.kumar.ec@gmail.com", "42", "Electronics", "2"),
                ("Abhay Thakur", "abhay.thakur.ec@gmail.com", "43", "Electronics", "2"),
                (
                    "Parshant Sharma",
                    "parshant.sharma.ec@gmail.com",
                    "44",
                    "Electronics",
                    "2",
                ),
                ("Uday Thakur", "uday.thakur.ec@gmail.com", "45", "Electronics", "2"),
                ("Anshika", "anshika.ec@gmail.com", "46", "Electronics", "2"),
                ("Vansh Rana", "vansh.rana.ec@gmail.com", "47", "Electronics", "2"),
                (
                    "Disha Choudhary",
                    "disha.choudhary.ec@gmail.com",
                    "48",
                    "Electronics",
                    "2",
                ),
                ("Mannat Walia", "mannat.walia.ec@gmail.com", "49", "Electronics", "2"),
                (
                    "Krish Singh Athwal",
                    "krish.singh.athwal.ec@gmail.com",
                    "50",
                    "Electronics",
                    "2",
                ),
                # Instrumentation - 2nd Semester (40 students)
                (
                    "Abhishek Sharma",
                    "abhishek.sharma.in@gmail.com",
                    "1",
                    "Instrumentation",
                    "2",
                ),
                ("Aditya", "aditya.in@gmail.com", "2", "Instrumentation", "2"),
                (
                    "Aditya Rana",
                    "aditya.rana.in@gmail.com",
                    "3",
                    "Instrumentation",
                    "2",
                ),
                ("Aashay Kumar", "aashay@gmail.com", "4", "Instrumentation", "2"),
                ("Anamika Lagwal", "anamika@gmail.com", "5", "Instrumentation", "2"),
                ("Ankit", "ankit.in@gmail.com", "6", "Instrumentation", "2"),
                ("Anjana Devi", "anjana@gmail.com", "7", "Instrumentation", "2"),
                ("Anusha", "anusha@gmail.com", "8", "Instrumentation", "2"),
                ("Arpit Koundal", "arpit@gmail.com", "9", "Instrumentation", "2"),
                (
                    "Aryan Kumar",
                    "aryan.kumar.in@gmail.com",
                    "10",
                    "Instrumentation",
                    "2",
                ),
                (
                    "Ashish Choudhary",
                    "ashish.choudhary.in@gmail.com",
                    "11",
                    "Instrumentation",
                    "2",
                ),
                ("Ashil", "ashil.in@gmail.com", "12", "Instrumentation", "2"),
                ("Ayush", "ayush.in@gmail.com", "13", "Instrumentation", "2"),
                ("Jayesh", "jayesh@gmail.com", "14", "Instrumentation", "2"),
                ("Kaashish", "kaashish@gmail.com", "15", "Instrumentation", "2"),
                (
                    "Krish Koundal",
                    "krish.koundal.in@gmail.com",
                    "16",
                    "Instrumentation",
                    "2",
                ),
                (
                    "Manoj Kumar",
                    "manoj.kumar.in@gmail.com",
                    "17",
                    "Instrumentation",
                    "2",
                ),
                ("Palak", "palak.in@gmail.com", "18", "Instrumentation", "2"),
                (
                    "Rahul Choudhary",
                    "rahul.choudhary.in@gmail.com",
                    "19",
                    "Instrumentation",
                    "2",
                ),
                (
                    "Rahul Maharaja",
                    "rahul.maharaja@gmail.com",
                    "20",
                    "Instrumentation",
                    "2",
                ),
                ("Reetul Choudhary", "reetul@gmail.com", "21", "Instrumentation", "2"),
                ("Ritam", "ritam@gmail.com", "22", "Instrumentation", "2"),
                ("Sahil", "sahil.in@gmail.com", "23", "Instrumentation", "2"),
                (
                    "Shardham Kumar",
                    "shardham.in2@gmail.com",
                    "24",
                    "Instrumentation",
                    "2",
                ),
                ("Sourav Dogra", "sourav@gmail.com", "25", "Instrumentation", "2"),
                ("Sujal", "sujal.in@gmail.com", "26", "Instrumentation", "2"),
                ("Sujan Kumar", "sujan@gmail.com", "27", "Instrumentation", "2"),
                ("Tulisha", "tulisha@gmail.com", "28", "Instrumentation", "2"),
                ("Tushar Kumar", "tushar@gmail.com", "29", "Instrumentation", "2"),
                ("Vinod Prasad", "vinod@gmail.com", "30", "Instrumentation", "2"),
                ("Aditya", "aditya.in2@gmail.com", "31", "Instrumentation", "2"),
                ("Anu Rana", "anu@gmail.com", "32", "Instrumentation", "2"),
                ("Bhata", "bhata@gmail.com", "33", "Instrumentation", "2"),
                (
                    "Piyush Kumar",
                    "piyush.kumar.in@gmail.com",
                    "34",
                    "Instrumentation",
                    "2",
                ),
                ("Prince Kumar", "prince@gmail.com", "35", "Instrumentation", "2"),
                ("Shavnam", "shavnam@gmail.com", "36", "Instrumentation", "2"),
                ("Shivam", "shivam.in@gmail.com", "37", "Instrumentation", "2"),
                (
                    "Sunil Kumar",
                    "sunil.kumar.in@gmail.com",
                    "38",
                    "Instrumentation",
                    "2",
                ),
                (
                    "Sujal Kumar",
                    "sujal.kumar.in@gmail.com",
                    "39",
                    "Instrumentation",
                    "2",
                ),
                (
                    "Sumit Kumar",
                    "sumit.kumar.in@gmail.com",
                    "40",
                    "Instrumentation",
                    "2",
                ),
                # Instrumentation - 4th Semester (44 students)
                ("Ajay", "ajay@gmail.com", "1", "Instrumentation", "4"),
                ("Akhil", "akhil.in4@gmail.com", "2", "Instrumentation", "4"),
                ("Akshay Rana", "akshay.rana@gmail.com", "3", "Instrumentation", "4"),
                ("Amandeep Singh", "amandeep@gmail.com", "4", "Instrumentation", "4"),
                ("Aman Joshi", "aman.joshi@gmail.com", "5", "Instrumentation", "4"),
                ("Amrit Kumar", "amrit@gmail.com", "6", "Instrumentation", "4"),
                (
                    "Ananya Choudhary",
                    "ananya.choudhary@gmail.com",
                    "7",
                    "Instrumentation",
                    "4",
                ),
                ("Ankit", "ankit.in4@gmail.com", "8", "Instrumentation", "4"),
                ("Anmol Verma", "anmol.verma@gmail.com", "9", "Instrumentation", "4"),
                ("Anshul", "anshul.in4@gmail.com", "10", "Instrumentation", "4"),
                (
                    "Arjun Kumar",
                    "arjun.kumar.in@gmail.com",
                    "11",
                    "Instrumentation",
                    "4",
                ),
                (
                    "Arjun Koundal",
                    "arjun.koundal@gmail.com",
                    "12",
                    "Instrumentation",
                    "4",
                ),
                ("Arun", "arun.in@gmail.com", "13", "Instrumentation", "4"),
                ("Arun Rana", "arun.rana@gmail.com", "14", "Instrumentation", "4"),
                ("Ashu Sharma", "ashu.sharma@gmail.com", "15", "Instrumentation", "4"),
                ("Ashish", "ashish.in4@gmail.com", "16", "Instrumentation", "4"),
                ("Ashok", "ashok.in@gmail.com", "17", "Instrumentation", "4"),
                ("Asmita", "asmita@gmail.com", "18", "Instrumentation", "4"),
                ("Aswani", "aswani@gmail.com", "19", "Instrumentation", "4"),
                (
                    "Aswani Kumar",
                    "aswani.kumar@gmail.com",
                    "20",
                    "Instrumentation",
                    "4",
                ),
                (
                    "Atish Sharma",
                    "atish.sharma@gmail.com",
                    "21",
                    "Instrumentation",
                    "4",
                ),
                ("Avnish", "avnish.in@gmail.com", "22", "Instrumentation", "4"),
                ("Ayush", "ayush.in4@gmail.com", "23", "Instrumentation", "4"),
                ("Ayush Raj", "ayush.raj@gmail.com", "24", "Instrumentation", "4"),
                ("Aditya", "aditya.in4@gmail.com", "25", "Instrumentation", "4"),
                (
                    "Aditya Sharma",
                    "aditya.sharma.in@gmail.com",
                    "26",
                    "Instrumentation",
                    "4",
                ),
                ("Aman", "aman.in@gmail.com", "27", "Instrumentation", "4"),
                (
                    "Aman Choudhary",
                    "aman.choudhary.in@gmail.com",
                    "28",
                    "Instrumentation",
                    "4",
                ),
                ("Apoorva", "apoorva.in@gmail.com", "29", "Instrumentation", "4"),
                ("Aakash Kumar", "aakash@gmail.com", "30", "Instrumentation", "4"),
                ("Aakul", "aakul@gmail.com", "31", "Instrumentation", "4"),
                (
                    "Aakul Vasarla",
                    "aakul.vasarla@gmail.com",
                    "32",
                    "Instrumentation",
                    "4",
                ),
                ("Abhashit", "abhashit@gmail.com", "33", "Instrumentation", "4"),
                ("Abhay", "abhay.in@gmail.com", "34", "Instrumentation", "4"),
                (
                    "Abhishek Kumar",
                    "abhishek.kumar.in@gmail.com",
                    "35",
                    "Instrumentation",
                    "4",
                ),
                ("Arjun", "arjun.in4@gmail.com", "36", "Instrumentation", "4"),
                ("Aryan", "aryan.in@gmail.com", "37", "Instrumentation", "4"),
                ("Arun Sharma", "arun.sharma@gmail.com", "38", "Instrumentation", "4"),
                ("Angad", "angad@gmail.com", "39", "Instrumentation", "4"),
                ("Ankita", "ankita.in@gmail.com", "40", "Instrumentation", "4"),
                (
                    "Anushka Sharma",
                    "anushka.sharma.in@gmail.com",
                    "41",
                    "Instrumentation",
                    "4",
                ),
                ("Rishant Agarwal", "rishant@gmail.com", "42", "Instrumentation", "4"),
                (
                    "Rajesh Kumar",
                    "rajesh.kumar.in@gmail.com",
                    "43",
                    "Instrumentation",
                    "4",
                ),
                ("Raj Kumar", "raj.kumar.in@gmail.com", "44", "Instrumentation", "4"),
                # Instrumentation - 6th Semester (as per notice)
                (
                    "Abhishek Kumar",
                    "abhishek.kumar.in6@gmail.com",
                    "1",
                    "Instrumentation",
                    "6",
                ),
                ("Akshay", "akshay.in6@gmail.com", "2", "Instrumentation", "6"),
                (
                    "Animesh Rana",
                    "animesh.rana.in6@gmail.com",
                    "3",
                    "Instrumentation",
                    "6",
                ),
                ("Anjali", "anjali.in6@gmail.com", "4", "Instrumentation", "6"),
                ("Anu Kumari", "anu.kumari.in6@gmail.com", "5", "Instrumentation", "6"),
                ("Ayush", "ayush1.in6@gmail.com", "6", "Instrumentation", "6"),
                ("Ayush", "ayush2.in6@gmail.com", "7", "Instrumentation", "6"),
                (
                    "Gaurav Shukla",
                    "gaurav.shukla.in6@gmail.com",
                    "8",
                    "Instrumentation",
                    "6",
                ),
                ("Harshit", "harshit.in6@gmail.com", "9", "Instrumentation", "6"),
                (
                    "Keshav (TFW)",
                    "keshav.tfw.in6@gmail.com",
                    "10",
                    "Instrumentation",
                    "6",
                ),
                ("Neha", "neha.in6@gmail.com", "11", "Instrumentation", "6"),
                (
                    "Nikhil Choudhary",
                    "nikhil.choudhary.in6@gmail.com",
                    "12",
                    "Instrumentation",
                    "6",
                ),
                (
                    "Pratham Choudhary",
                    "pratham.choudhary.in6@gmail.com",
                    "13",
                    "Instrumentation",
                    "6",
                ),
                ("Priya", "priya.in6@gmail.com", "14", "Instrumentation", "6"),
                ("Priyanshu", "priyanshu.in6@gmail.com", "15", "Instrumentation", "6"),
                (
                    "Priyanshu Choudhary",
                    "priyanshu.choudhary.in6@gmail.com",
                    "16",
                    "Instrumentation",
                    "6",
                ),
                ("Rithik", "rithik.in6@gmail.com", "17", "Instrumentation", "6"),
                (
                    "Sahib Singh Bedi",
                    "sahib.singh.bedi.in6@gmail.com",
                    "18",
                    "Instrumentation",
                    "6",
                ),
                ("Sahil", "sahil.in6@gmail.com", "19", "Instrumentation", "6"),
                ("Shahil", "shahil.in6@gmail.com", "20", "Instrumentation", "6"),
                ("Shubham", "shubham1.in6@gmail.com", "21", "Instrumentation", "6"),
                ("Sourabh", "sourabh.in6@gmail.com", "22", "Instrumentation", "6"),
                (
                    "Varun (TFW)",
                    "varun.tfw.in6@gmail.com",
                    "23",
                    "Instrumentation",
                    "6",
                ),
                ("Varun", "varun.in6@gmail.com", "24", "Instrumentation", "6"),
                (
                    "Vinay Kumar",
                    "vinay.kumar.in6@gmail.com",
                    "25",
                    "Instrumentation",
                    "6",
                ),
                ("Shubham", "shubham2.in6@gmail.com", "26", "Instrumentation", "6"),
                ("Akanksha", "akanksha.in6@gmail.com", "27", "Instrumentation", "6"),
                (
                    "Aman Kumar",
                    "aman.kumar.in6@gmail.com",
                    "28",
                    "Instrumentation",
                    "6",
                ),
                (
                    "Aryan Choudhary",
                    "aryan.choudhary.in6@gmail.com",
                    "29",
                    "Instrumentation",
                    "6",
                ),
                ("Dishant", "dishant.in6@gmail.com", "30", "Instrumentation", "6"),
                (
                    "Madhu Bala",
                    "madhu.bala.in6@gmail.com",
                    "31",
                    "Instrumentation",
                    "6",
                ),
                ("Shivam", "shivam.in6@gmail.com", "32", "Instrumentation", "6"),
                ("Prince", "prince.in6@gmail.com", "33", "Instrumentation", "6"),
                ("Abhishek", "abhishek.in6@gmail.com", "34", "Instrumentation", "6"),
                ("Akshay", "akshay2.in6@gmail.com", "35", "Instrumentation", "6"),
                (
                    "Rijul Kalia",
                    "rijul.kalia.in6@gmail.com",
                    "36",
                    "Instrumentation",
                    "6",
                ),
                # Mechanical - Semester 6
                (
                    "Akshit Pathania",
                    "akshit.pathania.me6@gmail.com",
                    "01",
                    "Mechanical",
                    "6",
                ),
                ("Aman Kumar", "aman.kumar.me6@gmail.com", "02", "Mechanical", "6"),
                ("Aman Sanjit", "aman.sanjit.me6@gmail.com", "03/2", "Mechanical", "6"),
                (
                    "Anirudh Kumar",
                    "anirudh.kumar.me6@gmail.com",
                    "04",
                    "Mechanical",
                    "6",
                ),
                ("Anmol", "anmol.me6@gmail.com", "05", "Mechanical", "6"),
                ("Anshul Kumar", "anshul.kumar.me6@gmail.com", "06", "Mechanical", "6"),
                ("Anshul Rana", "anshul.rana.me6@gmail.com", "07", "Mechanical", "6"),
                ("Anshul Rana", "anshul.rana2.me6@gmail.com", "08", "Mechanical", "6"),
                ("Archit Kumar", "archit.kumar.me6@gmail.com", "09", "Mechanical", "6"),
                ("Arpit", "arpit.me6@gmail.com", "10", "Mechanical", "6"),
                ("Aryan", "aryan.me6@gmail.com", "11", "Mechanical", "6"),
                (
                    "Chandan Kumar",
                    "chandan.kumar.me6@gmail.com",
                    "12",
                    "Mechanical",
                    "6",
                ),
                ("Deepak", "deepak.me6@gmail.com", "13", "Mechanical", "6"),
                (
                    "Divyansh Thakur",
                    "divyansh.thakur.me6@gmail.com",
                    "14",
                    "Mechanical",
                    "6",
                ),
                ("Freed Ali", "freed.ali.me6@gmail.com", "15", "Mechanical", "6"),
                ("Jasman", "jasman.me6@gmail.com", "16", "Mechanical", "6"),
                (
                    "Mandeep Singh",
                    "mandeep.singh.me6@gmail.com",
                    "17",
                    "Mechanical",
                    "6",
                ),
                ("Nikhil", "nikhil.me6@gmail.com", "18", "Mechanical", "6"),
                (
                    "Nikhil Choudhary",
                    "nikhil.choudhary.me6@gmail.com",
                    "19",
                    "Mechanical",
                    "6",
                ),
                ("Nishant", "nishant.me6@gmail.com", "20", "Mechanical", "6"),
                ("Nitish Kumar", "nitish.kumar.me6@gmail.com", "21", "Mechanical", "6"),
                (
                    "Pardeep Kumar",
                    "pardeep.kumar.me6@gmail.com",
                    "22",
                    "Mechanical",
                    "6",
                ),
                ("Piyush", "piyush.me6@gmail.com", "23", "Mechanical", "6"),
                (
                    "Rahul Koundal",
                    "rahul.koundal.me6@gmail.com",
                    "24",
                    "Mechanical",
                    "6",
                ),
                (
                    "Rajnish Choudhary",
                    "rajnish.choudhary.me6@gmail.com",
                    "25",
                    "Mechanical",
                    "6",
                ),
                ("Rajul Parmar", "rajul.parmar.me6@gmail.com", "26", "Mechanical", "6"),
                ("Raman Kumar", "raman.kumar.me6@gmail.com", "27", "Mechanical", "6"),
                ("Sahil Goma", "sahil.goma.me6@gmail.com", "28", "Mechanical", "6"),
                ("Sandeep Rana", "sandeep.rana.me6@gmail.com", "29", "Mechanical", "6"),
                (
                    "Shagun Singh Pathania",
                    "shagun.pathania.me6@gmail.com",
                    "30",
                    "Mechanical",
                    "6",
                ),
                (
                    "Subham Kondal",
                    "subham.kondal.me6@gmail.com",
                    "31",
                    "Mechanical",
                    "6",
                ),
                ("Suraj Kumar", "suraj.kumar.me6@gmail.com", "32", "Mechanical", "6"),
                ("Tarun Pal", "tarun.pal.me6@gmail.com", "33", "Mechanical", "6"),
                (
                    "Vanshul Sharma",
                    "vanshul.sharma.me6@gmail.com",
                    "34",
                    "Mechanical",
                    "6",
                ),
                ("Vishal Kumar", "vishal.kumar.me6@gmail.com", "35", "Mechanical", "6"),
                (
                    "Rupansh Koundal",
                    "rupansh.koundal.me6@gmail.com",
                    "36",
                    "Mechanical",
                    "6",
                ),
                (
                    "Abhishek Bhardwaj",
                    "abhishek.me6@gmail.com",
                    "37",
                    "Mechanical",
                    "6",
                ),
                ("Ansul Kumar", "ansu.kumar.me6@gmail.com", "38", "Mechanical", "6"),
                (
                    "Prince Bhatia",
                    "prince.bhatia.me6@gmail.com",
                    "39",
                    "Mechanical",
                    "6",
                ),
                (
                    "Sourav Tatyal",
                    "sourav.tatyal.me6@gmail.com",
                    "40",
                    "Mechanical",
                    "6",
                ),
                ("Vivek", "vivek.me6@gmail.com", "41", "Mechanical", "6"),
                (
                    "Thakur Nitish Singh Baryal",
                    "nitish.baryal.me6@gmail.com",
                    "42",
                    "Mechanical",
                    "6",
                ),
                ("Akshit", "akshit.me6@gmail.com", "43", "Mechanical", "6"),
                ("Anurag Soni", "anurag.soni.me6@gmail.com", "44", "Mechanical", "6"),
            ]

            students = []
            for name, email, roll_num, branch, semester in students_data:
                student = User(
                    name=name,
                    email=email,
                    password=generate_password_hash(roll_num),
                    role="student",
                    roll_number=roll_num,
                    branch=branch,
                    semester=semester,
                )
                students.append(student)

            teacher_users = []
            for name, email, department, _subject in teachers_data:
                teacher_users.append(
                    User(
                        name=name,
                        email=email.lower(),
                        password=generate_password_hash("teacher123"),
                        role="teacher",
                        department=department,
                    )
                )

            db.session.add_all([admin, default_teacher] + teacher_users + students)
            db.session.commit()

    app.run(debug=True)
